from __future__ import annotations

from pathlib import Path
import json

import pandas as pd

VERSION = "COMMITTEE_EXECUTIVE_VIEWS_V1"
ROLE = "REPORTING_ONLY"
ACTION_HORIZONS = ("CT", "MT", "LT", "SHORT", "TOP_DOWN")
ETF_HORIZONS = ("CT", "MT", "LT", "SHORT", "TOP_DOWN")


def _read(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path, sep=";", encoding="utf-8-sig", low_memory=False)


def _numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def _series(frame: pd.DataFrame, field: str, default=False) -> pd.Series:
    if field in frame.columns:
        return frame[field]
    return pd.Series(default, index=frame.index)


def _bool_series(frame: pd.DataFrame, field: str) -> pd.Series:
    values = _series(frame, field, False)
    if pd.api.types.is_bool_dtype(values):
        return values.fillna(False).astype(bool)
    return values.astype(str).str.strip().str.lower().isin({"true", "1", "yes", "oui"})


def _priority(decision: object) -> int:
    value = str(decision or "").upper()
    order = {
        "BUY_CANDIDATE": 0,
        "FAVORABLE": 0,
        "SHORT_RISK_CANDIDATE": 0,
        "WATCH": 1,
        "WATCH_SHORT_RISK": 1,
        "REVIEW": 2,
        "NEUTRAL": 2,
        "REJECT": 3,
        "NO_SHORT_RISK": 3,
        "DEFAVORABLE": 3,
    }
    if value.startswith("BLOCK") or value.startswith("SHADOW_BASELINE_REQUIRED"):
        return 9
    return order.get(value, 5)


def _rank_within(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    if not {"asset_class", "horizon"}.issubset(out.columns):
        return out
    out["score_num"] = _numeric(_series(out, "score", None))
    rank = pd.Series(pd.NA, index=out.index, dtype="Int64")
    for (_asset, _horizon), idx in out.groupby(["asset_class", "horizon"], dropna=False).groups.items():
        scored = out.loc[idx, "score_num"]
        valid = scored.notna()
        if valid.any():
            rank.loc[scored.index[valid]] = scored.loc[valid].rank(method="first", ascending=False).astype("Int64")
    out["committee_rank"] = rank
    return out


def build_action_priority(decisions: pd.DataFrame, top_n_per_horizon: int = 30) -> pd.DataFrame:
    if decisions.empty:
        return pd.DataFrame()
    ranked = _rank_within(decisions)
    required = {"asset_class", "horizon", "decision"}
    if not required.issubset(ranked.columns):
        return pd.DataFrame()
    ranked = ranked[
        ranked["asset_class"].astype(str).eq("ACTION")
        & ranked["horizon"].astype(str).isin(ACTION_HORIZONS)
    ].copy()
    if ranked.empty:
        return pd.DataFrame()
    ranked["decision_priority"] = ranked["decision"].map(_priority)
    ranked["coverage_num"] = _numeric(_series(ranked, "coverage_pct", None))
    for field in ("name", "isin"):
        if field not in ranked.columns:
            ranked[field] = ""
    ranked = ranked.sort_values(
        ["horizon", "decision_priority", "score_num", "coverage_num", "name", "isin"],
        ascending=[True, True, False, False, True, True],
        na_position="last",
    )
    ranked = ranked.groupby("horizon", group_keys=False).head(top_n_per_horizon).reset_index(drop=True)
    keep = [
        "horizon", "committee_rank", "isin", "name", "sector", "score", "coverage_pct", "status", "decision",
        "score_source", "backtest_attribution", "action_reference_score", "action_reference_decision",
        "action_challenger_score", "action_challenger_decision", "action_score_delta_challenger_vs_reference",
    ]
    return ranked[[c for c in keep if c in ranked.columns]]


def build_etf_top30(decisions: pd.DataFrame, top_n: int = 30) -> pd.DataFrame:
    if decisions.empty:
        return pd.DataFrame()
    ranked = _rank_within(decisions)
    if not {"asset_class", "horizon", "isin"}.issubset(ranked.columns):
        return pd.DataFrame()
    etf = ranked[
        ranked["asset_class"].astype(str).eq("ETF")
        & ranked["horizon"].astype(str).isin(ETF_HORIZONS)
    ].copy()
    if etf.empty:
        return pd.DataFrame()
    for field in ("name", "sector"):
        if field not in etf.columns:
            etf[field] = ""
    base = etf[["isin", "name", "sector"]].drop_duplicates("isin", keep="first").copy()
    for horizon in ETF_HORIZONS:
        sub = etf[etf["horizon"].astype(str).eq(horizon)].copy()
        if sub.empty:
            continue
        rename = {
            "score": f"score_{horizon.lower()}",
            "coverage_pct": f"coverage_{horizon.lower()}_pct",
            "status": f"status_{horizon.lower()}",
            "decision": f"decision_{horizon.lower()}",
            "committee_rank": f"rank_{horizon.lower()}",
            "score_source": f"source_{horizon.lower()}",
            "backtest_attribution": f"backtest_attribution_{horizon.lower()}",
        }
        cols = ["isin"] + [c for c in rename if c in sub.columns]
        base = base.merge(sub[cols].rename(columns=rename), on="isin", how="left")

    def selected(row: pd.Series, horizon: str) -> bool:
        value = str(row.get(f"decision_{horizon.lower()}", "")).upper()
        if horizon == "TOP_DOWN":
            return value == "FAVORABLE"
        if horizon == "SHORT":
            return value == "SHORT_RISK_CANDIDATE"
        return value == "BUY_CANDIDATE"

    for horizon in ETF_HORIZONS:
        base[f"selected_{horizon.lower()}"] = base.apply(lambda row, h=horizon: selected(row, h), axis=1)
    flag_cols = [f"selected_{h.lower()}" for h in ETF_HORIZONS]
    base["selected_horizon_count"] = base[flag_cols].sum(axis=1)
    base["mt_reference_validated_flag"] = _series(base, "source_mt", "").astype(str).str.contains("V20.8.1", regex=False)
    sort_cols = [c for c in ("rank_mt", "rank_ct", "rank_lt", "name") if c in base.columns]
    if sort_cols:
        base = base.sort_values(sort_cols, ascending=[True] * len(sort_cols), na_position="last")
    return base.head(top_n).reset_index(drop=True)


def _first_numeric(frame: pd.DataFrame, fields: tuple[str, ...]) -> pd.Series | None:
    for field in fields:
        if field in frame.columns:
            values = _numeric(frame[field])
            if values.notna().any():
                return values
    return None


def enrich_tct_details(details: pd.DataFrame, baseline: pd.DataFrame) -> pd.DataFrame:
    if details.empty or baseline.empty or "isin" not in baseline.columns:
        return details.copy()
    out = details.copy()
    wanted = [
        "isin", "market_cap", "market_cap_yf", "rvol20", "rvol20_3d_avg", "volume", "currency", "trading_currency",
        "earnings_date_finnhub", "earnings_hour_finnhub", "eps_estimate_analysts_finnhub",
        "eps_estimate_dispersion_pct_finnhub", "amf_public_short_holders_count",
        "amf_public_short_max_holder_pct", "amf_short_data_as_of",
    ]
    cols = [c for c in wanted if c in baseline.columns]
    if "isin" not in cols:
        return out
    source = baseline[cols].drop_duplicates("isin", keep="last").copy()
    out = out.merge(source, on="isin", how="left", suffixes=("", "_baseline"))
    market_cap = _first_numeric(out, ("market_cap", "market_cap_yf"))
    if market_cap is not None:
        out["market_cap_native_m"] = (market_cap / 1_000_000.0).round(2)
    if "rvol20" in out.columns:
        out["relative_volume_20d"] = _numeric(out["rvol20"]).round(3)
    return out


def _top_names(group: pd.DataFrame, n: int = 3) -> str:
    work = group.copy()
    work["_rank"] = _numeric(_series(work, "tct_baseline_rank", None))
    work["_score"] = _numeric(_series(work, "tct_baseline_score", None))
    if "name" not in work.columns:
        work["name"] = ""
    work = work.sort_values(["_rank", "_score", "name"], ascending=[True, False, True], na_position="last")
    names = [str(v).strip() for v in work["name"].head(n) if str(v).strip()]
    return " | ".join(names)


def build_tct_views(details: pd.DataFrame, dashboard: pd.DataFrame, baseline: pd.DataFrame) -> dict[str, pd.DataFrame]:
    if details.empty:
        return {
            "top50": pd.DataFrame(), "t1": pd.DataFrame(), "t2": pd.DataFrame(),
            "earnings_d0_5": pd.DataFrame(), "event_risk": pd.DataFrame(), "sectors": dashboard.copy(),
        }
    d = enrich_tct_details(details, baseline)
    d["rank_num"] = _numeric(_series(d, "tct_baseline_rank", None))
    d["days_num"] = _numeric(_series(d, "days_to_earnings", None))
    d["timing_quality_num"] = _numeric(_series(d, "timing_quality_score", None))
    if "isin" not in d.columns:
        d["isin"] = ""

    top50 = d[d["rank_num"].notna()].sort_values(["rank_num", "isin"]).head(50).copy()
    t1 = d[_bool_series(d, "timing_t1_flag")].sort_values(
        ["rank_num", "timing_quality_num"], ascending=[True, False], na_position="last"
    ).copy()
    t2 = d[_bool_series(d, "timing_t2_flag")].sort_values(
        ["rank_num", "timing_quality_num"], ascending=[True, False], na_position="last"
    ).copy()
    earnings = d[_series(d, "earnings_bucket", "").astype(str).isin({"EARNINGS_D0_1", "EARNINGS_D2_5"})].sort_values(
        ["days_num", "rank_num"], na_position="last"
    ).copy()
    event_risk = d[_bool_series(d, "event_gap_risk_flag")].sort_values(["rank_num", "isin"], na_position="last").copy()

    sectors = dashboard.copy()
    if not sectors.empty and "sector" in sectors.columns and "sector" in d.columns:
        top_map = {str(sector): _top_names(group) for sector, group in d.groupby("sector", dropna=False)}
        sectors["top_3_baseline"] = sectors["sector"].astype(str).map(top_map).fillna("")
        signal_map = {}
        signal_rows = d[_bool_series(d, "timing_t1_flag") | _bool_series(d, "timing_t2_flag")]
        for sector, group in signal_rows.groupby("sector", dropna=False):
            signal_map[str(sector)] = _top_names(group)
        sectors["top_3_timing_shadow"] = sectors["sector"].astype(str).map(signal_map).fillna("")

    presentation_cols = [
        "isin", "name", "sector", "last_price", "price_currency", "market_cap_native_m", "relative_volume_20d",
        "tct_baseline_rank", "tct_baseline_score", "tct_baseline_coverage_pct", "tct_baseline_status",
        "timing_setup", "timing_decision", "timing_status", "timing_quality_score", "t1_quality_score", "t2_quality_score",
        "days_to_earnings", "earnings_date_finnhub", "earnings_hour_finnhub", "earnings_bucket", "event_risk_level",
        "event_gap_risk_flag", "eps_revision_3m", "eps_estimate_analysts_finnhub", "eps_estimate_dispersion_pct_finnhub",
        "short_interest_pct", "amf_public_short_holders_count", "amf_public_short_max_holder_pct", "amf_short_data_as_of",
        "sector_classification_quality", "sector_gap_flag", "timing_rejection_reason", "reporting_role", "t1_t2_score_influence",
        "live_orders_enabled",
    ]

    def present(frame: pd.DataFrame) -> pd.DataFrame:
        return frame[[c for c in presentation_cols if c in frame.columns]].reset_index(drop=True)

    return {
        "top50": present(top50),
        "t1": present(t1),
        "t2": present(t2),
        "earnings_d0_5": present(earnings),
        "event_risk": present(event_risk),
        "sectors": sectors.reset_index(drop=True),
    }


def build_quality_summary(decisions: pd.DataFrame, tct_details: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if not decisions.empty and {"asset_class", "horizon"}.issubset(decisions.columns):
        work = decisions.copy()
        work["coverage_num"] = _numeric(_series(work, "coverage_pct", None))
        for (asset, horizon), group in work.groupby(["asset_class", "horizon"], dropna=False):
            blocked = group["status"].astype(str).str.startswith("BLOCK").sum() if "status" in group.columns else 0
            rows.append({
                "scope": f"{asset}:{horizon}",
                "rows": int(len(group)),
                "mean_coverage_pct": round(float(group["coverage_num"].mean()), 2) if group["coverage_num"].notna().any() else None,
                "min_coverage_pct": round(float(group["coverage_num"].min()), 2) if group["coverage_num"].notna().any() else None,
                "blocked_rows": int(blocked),
                "data_gap_rows": int(group["coverage_num"].isna().sum()),
                "note": "Committee final decision coverage; no neutral missing-data imputation.",
            })
    if not tct_details.empty:
        coverage = _numeric(_series(tct_details, "tct_baseline_coverage_pct", None))
        rows.append({
            "scope": "ACTION:TCT_CONTEXT",
            "rows": int(len(tct_details)),
            "mean_coverage_pct": round(float(coverage.mean()), 2) if coverage.notna().any() else None,
            "min_coverage_pct": round(float(coverage.min()), 2) if coverage.notna().any() else None,
            "blocked_rows": 0,
            "data_gap_rows": int(_bool_series(tct_details, "timing_data_gap_flag").sum()),
            "note": "TCT context only; T1/T2 influence remains zero and live execution remains forbidden.",
        })
    return pd.DataFrame(rows)


def _style_workbook(path: Path) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in ws.columns:
            values = [str(c.value) if c.value is not None else "" for c in list(col)[:80]]
            width = min(max(max((len(v) for v in values), default=0) + 2, 10), 34)
            ws.column_dimensions[col[0].column_letter].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def write_outputs(
    action_priority: pd.DataFrame,
    etf_top30: pd.DataFrame,
    tct_views: dict[str, pd.DataFrame],
    quality: pd.DataFrame,
    decisions: pd.DataFrame,
    outdir: Path,
) -> dict:
    outdir.mkdir(parents=True, exist_ok=True)
    action_csv = outdir / "ACTION_COMMITTEE_PRIORITY_BY_HORIZON.csv"
    etf_csv = outdir / "ETF_COMMITTEE_TOP30.csv"
    tct_csv = outdir / "TCT_COMMITTEE_TOP50.csv"
    earnings_csv = outdir / "TCT_EARNINGS_D0_5_CONTEXT.csv"
    quality_csv = outdir / "COMMITTEE_DATA_QUALITY.csv"
    summary_json = outdir / "COMMITTEE_EXECUTIVE_VIEWS_SUMMARY.json"
    workbook = outdir / "COMMITTEE_EXECUTIVE_VIEWS.xlsx"

    action_priority.to_csv(action_csv, sep=";", index=False, encoding="utf-8-sig")
    etf_top30.to_csv(etf_csv, sep=";", index=False, encoding="utf-8-sig")
    tct_views["top50"].to_csv(tct_csv, sep=";", index=False, encoding="utf-8-sig")
    tct_views["earnings_d0_5"].to_csv(earnings_csv, sep=";", index=False, encoding="utf-8-sig")
    quality.to_csv(quality_csv, sep=";", index=False, encoding="utf-8-sig")

    decision_counts = []
    if not decisions.empty and {"asset_class", "horizon", "decision"}.issubset(decisions.columns):
        decision_counts = decisions.groupby(["asset_class", "horizon", "decision"], dropna=False).size().reset_index(name="count").to_dict("records")
    summary = {
        "version": VERSION,
        "role": ROLE,
        "action_priority_rows": int(len(action_priority)),
        "etf_top30_rows": int(len(etf_top30)),
        "tct_top50_rows": int(len(tct_views["top50"])),
        "tct_t1_shadow_rows": int(len(tct_views["t1"])),
        "tct_t2_shadow_rows": int(len(tct_views["t2"])),
        "tct_earnings_d0_5_rows": int(len(tct_views["earnings_d0_5"])),
        "tct_event_gap_risk_rows": int(len(tct_views["event_risk"])),
        "score_changes": False,
        "new_composite_opportunity_score": False,
        "fixed_probability_or_expectancy_added": False,
        "t1_t2_score_influence": 0.0,
        "live_orders_enabled": False,
        "decision_counts": decision_counts,
    }
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    counts = pd.DataFrame(decision_counts)
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        counts.to_excel(writer, sheet_name="Synthese", index=False)
        action_priority.to_excel(writer, sheet_name="Actions_Priorites", index=False)
        etf_top30.to_excel(writer, sheet_name="ETF_Top30", index=False)
        tct_views["top50"].to_excel(writer, sheet_name="TCT_Top50", index=False)
        tct_views["t1"].to_excel(writer, sheet_name="TCT_T1_Shadow", index=False)
        tct_views["t2"].to_excel(writer, sheet_name="TCT_T2_Shadow", index=False)
        tct_views["earnings_d0_5"].to_excel(writer, sheet_name="TCT_Earnings_D0_5", index=False)
        tct_views["event_risk"].to_excel(writer, sheet_name="TCT_Event_Risk", index=False)
        tct_views["sectors"].to_excel(writer, sheet_name="TCT_Secteurs", index=False)
        quality.to_excel(writer, sheet_name="Data_Quality", index=False)
    _style_workbook(workbook)

    return {
        "status": "SUCCESS",
        "version": VERSION,
        "role": ROLE,
        "workbook": str(workbook),
        "action_priority_csv": str(action_csv),
        "etf_top30_csv": str(etf_csv),
        "tct_top50_csv": str(tct_csv),
        "tct_earnings_d0_5_csv": str(earnings_csv),
        "quality_csv": str(quality_csv),
        "summary_json": str(summary_json),
        **summary,
    }


def run(root: Path) -> dict:
    outdir = root / "outputs" / "committee_master"
    decisions = _read(outdir / "COMMITTEE_DECISIONS.csv")
    if decisions.empty:
        raise FileNotFoundError("Missing or empty Committee decisions output")
    details = _read(outdir / "TCT_SECTOR_COMMITTEE_DETAILS.csv")
    dashboard = _read(outdir / "TCT_SECTOR_DASHBOARD.csv")
    baseline = _read(outdir / "TCT_BASELINE_V24_1_8.csv")
    action_priority = build_action_priority(decisions)
    etf_top30 = build_etf_top30(decisions)
    tct_views = build_tct_views(details, dashboard, baseline)
    quality = build_quality_summary(decisions, details)
    return write_outputs(action_priority, etf_top30, tct_views, quality, decisions, outdir)
