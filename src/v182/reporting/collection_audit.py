from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import os
import shutil
import threading
import pandas as pd

from v182.audit.provenance import actual_sources_by_field
from v182.io.frames import MISSING_TOKEN

SOURCE_HINTS = [
    (("morningstar",), "Morningstar / source autorisee ou snapshot attribue"),
    (("consensus", "target_", "upside", "upgrade", "downgrade", "analyst"), "Finnhub / yfinance / Boursorama / Zonebourse"),
    (("funnel_global_macro", "macro_", "real_yield", "inflation"), "FRED / BCE / Eurostat"),
    (("news", "bnis", "sentiment"), "AMF / Emetteur / GDELT / Finnhub"),
    (("ter", "aum", "tracking", "replication", "holdings"), "Emetteur/KID / Boursorama / Morningstar-authorized source"),
    (("dividend", "payout"), "yfinance / Emetteur / Boursorama"),
    (("sector", "industry", "country"), "yfinance / Euronext / referentiel"),
    (("rsi", "macd", "perf_", "mm", "volatility", "drawdown", "atr", "stoch", "rvol", "distance_high_52w", "catchup", "rotation"), "OHLCV yfinance -> calcul interne PIT"),
    (("per", "pb", "roe", "roa", "margin", "growth", "debt", "fcf", "market_cap"), "yfinance / Alpha Vantage / Emetteur"),
]
TECHNICAL_FIELDS={"canonical_seed_status"}
MISSING_TEXT_TOKENS={"", "MISSING", "UNKNOWN", MISSING_TOKEN, "NOT_LOADED", "NAN", "<NA>", "N/A", "NA", "NULL"}
STATE_UNCHANGED_WAVES={"WAVE_00_ETF_TICKERS","WAVE_01_ACTION_OHLCV","WAVE_02_ETF_OHLCV"}
_AUDIT_CACHE_LOCK=threading.RLock()
_LAST_INVENTORY: pd.DataFrame | None=None
_LAST_PROVENANCE: pd.DataFrame | None=None


def source_hint(field: str) -> str:
    low=str(field).lower()
    for tokens,source in SOURCE_HINTS:
        if any(token in low for token in tokens): return source
    return "Referentiel / source specifique du champ / non determinee"


def _missing_mask(series: pd.Series) -> pd.Series:
    """Vectorized equivalent of frames.is_missing for master/audit columns."""
    mask=series.isna()
    remaining=~mask
    if remaining.any():
        normalized=series.loc[remaining].astype(str).str.strip().str.upper()
        mask.loc[remaining]=normalized.isin(MISSING_TEXT_TOKENS)
    return mask.astype(bool)


def _field_status(frame: pd.DataFrame, asset_class: str, wave_id: str) -> pd.DataFrame:
    rows=[]; n=max(len(frame),1); identity={"isin","name"}|TECHNICAL_FIELDS
    generated=datetime.now(timezone.utc).isoformat()
    for field in frame.columns:
        if field in identity: continue
        available=int((~_missing_mask(frame[field])).sum()); coverage=available/n*100.0
        status="MISSING" if available==0 else "PARTIAL" if available<len(frame) else "AVAILABLE"
        rows.append({
            "collection":wave_id,"asset_class":asset_class,"field":field,"status":status,
            "available_rows":available,"missing_rows":int(len(frame)-available),"universe_rows":int(len(frame)),
            "coverage_pct":round(coverage,2),"source_theorique":source_hint(field),
            "generated_at_utc":generated,
        })
    return pd.DataFrame(rows)


def _format_excel(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb=load_workbook(path); header_fill=PatternFill("solid",fgColor="1F4E78")
    for ws in wb.worksheets:
        ws.freeze_panes="A2"; ws.sheet_view.showGridLines=False
        for cell in ws[1]:
            cell.font=Font(bold=True,color="FFFFFF"); cell.fill=header_fill; cell.alignment=Alignment(horizontal="center",vertical="center")
        for col in ws.columns:
            letter=col[0].column_letter; width=min(55,max(10,max((len(str(c.value)) if c.value is not None else 0) for c in col)+2)); ws.column_dimensions[letter].width=width
    wb.save(path)


def _github_compact_intermediate_enabled(wave_id: str) -> bool:
    """Use lossless CSV between waves on production GitHub runs; keep final XLSX."""
    if wave_id == "WAVE_99_FINAL":
        return False
    on_github=os.environ.get("GITHUB_ACTIONS","").strip().lower()=="true"
    source_mode=os.environ.get("PEA_SLOW_SOURCE_MODE","").strip().upper()
    return on_github and source_mode in {"LIVE","CACHE_PREFERRED"}


def _build_inventory(actions: pd.DataFrame, etfs: pd.DataFrame, wave_id: str) -> tuple[pd.DataFrame,pd.DataFrame]:
    inventory=pd.concat([_field_status(actions,"ACTION",wave_id),_field_status(etfs,"ETF",wave_id)],ignore_index=True)
    provenance=actual_sources_by_field()
    if not provenance.empty:
        provenance=provenance.rename(columns={"universe":"asset_class"})
        inventory=inventory.merge(provenance,on=["asset_class","field"],how="left")
    else:
        for col in ("sources_reelles","source_urls","evidence_levels","last_as_of"): inventory[col]=""
    for col in ("sources_reelles","source_urls","evidence_levels","last_as_of"):
        if col not in inventory.columns: inventory[col]=""
    inventory["source_reelle_absente"]=(inventory["sources_reelles"].fillna("").astype(str).str.strip()=="")
    return inventory,provenance


def _inventory_for_audit(
    actions: pd.DataFrame,
    etfs: pd.DataFrame,
    wave_id: str,
    *,
    reuse_previous_state: bool,
) -> tuple[pd.DataFrame,pd.DataFrame,bool]:
    """Reuse the previous full inventory only when the caller proves state unchanged."""
    global _LAST_INVENTORY,_LAST_PROVENANCE
    if reuse_previous_state:
        with _AUDIT_CACHE_LOCK:
            if _LAST_INVENTORY is not None and _LAST_PROVENANCE is not None:
                inventory=_LAST_INVENTORY.copy(deep=True)
                provenance=_LAST_PROVENANCE.copy(deep=True)
                generated=datetime.now(timezone.utc).isoformat()
                inventory["collection"]=wave_id
                inventory["generated_at_utc"]=generated
                return inventory,provenance,True

    inventory,provenance=_build_inventory(actions,etfs,wave_id)
    with _AUDIT_CACHE_LOCK:
        _LAST_INVENTORY=inventory.copy(deep=True)
        _LAST_PROVENANCE=provenance.copy(deep=True)
    return inventory,provenance,False


def _reset_audit_cache_for_tests() -> None:
    global _LAST_INVENTORY,_LAST_PROVENANCE
    with _AUDIT_CACHE_LOCK:
        _LAST_INVENTORY=None
        _LAST_PROVENANCE=None


def write_collection_audit(
    actions: pd.DataFrame, etfs: pd.DataFrame, wave_id: str, output_root: str | Path,
    *, failures: list[dict] | None = None, source_context: str = "", write_excel: bool = True,
    reuse_previous_state: bool = False,
) -> str:
    """Write a post-collection audit with retained-source provenance.

    GitHub production runs use compact lossless CSV between waves and still publish
    the final Excel audit once at WAVE_99. The three cache-only waves that provably
    cannot alter either master automatically reuse the preceding complete inventory;
    their own per-wave file, source context, failures and history remain published.
    All other waves always recompute unless an explicit trusted caller opts in.
    """
    root=Path(output_root); root.mkdir(parents=True,exist_ok=True)
    reuse=bool(reuse_previous_state or wave_id in STATE_UNCHANGED_WAVES)
    inventory,provenance,reused=_inventory_for_audit(actions,etfs,wave_id,reuse_previous_state=reuse)
    missing=inventory[inventory["status"]=="MISSING"].copy(); partial=inventory[inventory["status"]=="PARTIAL"].copy(); available=inventory[inventory["status"]=="AVAILABLE"].copy()
    summary=pd.DataFrame([
        {"collection":wave_id,"asset_class":"ACTION","universe_rows":len(actions),"missing_fields":int((inventory.query("asset_class=='ACTION' and status=='MISSING'")).shape[0]),"partial_fields":int((inventory.query("asset_class=='ACTION' and status=='PARTIAL'")).shape[0]),"fields_with_actual_source":int(((inventory.asset_class=="ACTION")&(~inventory.source_reelle_absente)).sum()),"source_context":source_context,"inventory_reused":reused},
        {"collection":wave_id,"asset_class":"ETF","universe_rows":len(etfs),"missing_fields":int((inventory.query("asset_class=='ETF' and status=='MISSING'")).shape[0]),"partial_fields":int((inventory.query("asset_class=='ETF' and status=='PARTIAL'")).shape[0]),"fields_with_actual_source":int(((inventory.asset_class=="ETF")&(~inventory.source_reelle_absente)).sum()),"source_context":source_context,"inventory_reused":reused},
    ])
    failures_df=pd.DataFrame(failures or []); safe="".join(ch if ch.isalnum() or ch in "_-" else "_" for ch in wave_id)[:40]
    compact_runtime=_github_compact_intermediate_enabled(wave_id)
    if compact_runtime:
        os.environ["PEA_EFFECTIVE_INTERMEDIATE_AUDIT_FORMAT"]="CSV"
    effective_write_excel=bool(write_excel) and not compact_runtime
    if effective_write_excel:
        path=root/f"COLLECTION_AUDIT_{safe}.xlsx"
        with pd.ExcelWriter(path,engine="openpyxl") as writer:
            summary.to_excel(writer,sheet_name="Synthese",index=False)
            missing.to_excel(writer,sheet_name="Donnees_non_disponibles",index=False)
            partial.to_excel(writer,sheet_name="Donnees_partielles",index=False)
            available.to_excel(writer,sheet_name="Donnees_disponibles",index=False)
            inventory[["field","asset_class","sources_reelles","source_urls","evidence_levels","last_as_of","source_theorique"]].drop_duplicates().to_excel(writer,sheet_name="Sources_reelles",index=False)
            provenance.to_excel(writer,sheet_name="Provenance_agregee",index=False)
            failures_df.to_excel(writer,sheet_name="Echecs_collecte",index=False)
        _format_excel(path); latest=root/"COLLECTION_DATA_AVAILABILITY_LATEST.xlsx"; shutil.copyfile(path,latest)
    else:
        path=root/f"COLLECTION_AUDIT_{safe}.csv"
        compact=inventory.copy(); compact["source_context"]=source_context; compact["inventory_reused"]=reused
        compact.to_csv(path,sep=";",encoding="utf-8-sig",index=False)
        if not failures_df.empty:
            failures_df.to_csv(root/f"COLLECTION_AUDIT_{safe}_FAILURES.csv",sep=";",encoding="utf-8-sig",index=False)
    history_path=root/"COLLECTION_AUDIT_HISTORY.csv"; hist=summary.copy(); hist["generated_at_utc"]=datetime.now(timezone.utc).isoformat(); hist.to_csv(history_path,sep=";",encoding="utf-8-sig",index=False,mode="a",header=not history_path.exists())
    return str(path)
