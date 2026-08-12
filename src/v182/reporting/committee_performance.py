from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import json
import math
import pandas as pd


def _read(path: Path, columns: list[str] | None=None) -> pd.DataFrame:
    if not path.exists(): return pd.DataFrame(columns=columns or [])
    try: return pd.read_csv(path,sep=";",encoding="utf-8-sig",low_memory=False)
    except Exception: return pd.DataFrame(columns=columns or [])


def _save(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True,exist_ok=True); df.to_csv(path,sep=";",encoding="utf-8-sig",index=False)


def _num(value):
    try:
        x=float(value); return x if math.isfinite(x) else None
    except (TypeError,ValueError): return None


def _sector(row: pd.Series) -> str:
    for field in ("sector_yf","sector_yahoo","sector","sector_bucket","industry_yf"):
        raw=row.get(field)
        if raw is not None and str(raw).strip().lower() not in {"","nan","none","unknown","n/a"}: return str(raw).strip()
    return "NON CLASSE"


def _price_master(actions: pd.DataFrame, etfs: pd.DataFrame) -> dict[str,dict]:
    out={}
    for asset,frame in (("ACTION",actions),("ETF",etfs)):
        if frame.empty or "isin" not in frame.columns: continue
        for _,row in frame.iterrows():
            price=None
            for field in ("last_close","current_price_yf","close"):
                price=_num(row.get(field))
                if price is not None and price>0: break
            out[str(row.get("isin","") or "")]={"price":price,"name":str(row.get("name","") or ""),"sector":_sector(row),"asset_class":asset}
    return out


def _style_xlsx(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment
    wb=load_workbook(path); dark="17365D"
    for ws in wb.worksheets:
        ws.freeze_panes="A2"; ws.sheet_view.showGridLines=False
        for cell in ws[1]: cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor=dark); cell.alignment=Alignment(horizontal="center")
        for col in ws.columns:
            letter=col[0].column_letter; ws.column_dimensions[letter].width=min(34,max(10,max(len(str(c.value)) if c.value is not None else 0 for c in col)+2))
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value,(int,float)): cell.number_format='#,##0.00;[Red](#,##0.00);-'
    if "NAV_Quotidienne" in wb.sheetnames:
        ws=wb["NAV_Quotidienne"]; headers={c.value:i+1 for i,c in enumerate(ws[1])}
        if ws.max_row>=3 and "date" in headers and "nav_eur" in headers:
            chart=LineChart(); chart.title="Performance cumulee du portefeuille virtuel"; chart.y_axis.title="NAV EUR"; chart.x_axis.title="Date"
            data=Reference(ws,min_col=headers["nav_eur"],min_row=1,max_row=ws.max_row); cats=Reference(ws,min_col=headers["date"],min_row=2,max_row=ws.max_row); chart.add_data(data,titles_from_data=True); chart.set_categories(cats); chart.height=8; chart.width=16; ws.add_chart(chart,"J2")
    wb.save(path)


def run(root: Path) -> dict:
    cfg=json.loads((root/"config"/"COMMITTEE_VIRTUAL_MONEY_MANAGEMENT.json").read_text(encoding="utf-8"))
    decisions_path=root/"outputs"/"committee_master"/"COMMITTEE_DECISIONS.csv"
    if not decisions_path.exists(): return {"status":"BLOCKED_NO_COMMITTEE_DECISIONS"}
    decisions=pd.read_csv(decisions_path,sep=";",encoding="utf-8-sig",low_memory=False)
    actions=_read(root/"outputs"/"V18.2_PEA_ACTIONS_MASTER_ENRICHED.csv"); etfs=_read(root/"outputs"/"V18.2_PEA_ETF_MASTER_ENRICHED.csv"); prices=_price_master(actions,etfs)
    state=root/"state"/"committee_performance"; state.mkdir(parents=True,exist_ok=True); outdir=root/"outputs"/"performance"; outdir.mkdir(parents=True,exist_ok=True)
    signals=_read(state/"signals.csv"); marks=_read(state/"daily_marks.csv"); positions=_read(state/"positions.csv"); tx=_read(state/"transactions.csv"); nav=_read(state/"nav.csv"); previous=_read(state/"last_decisions.csv")
    today=datetime.now(timezone.utc).date().isoformat(); buy_set=set(cfg["buy_decisions"]); min_buy_score=float(cfg.get("minimum_buy_score",77.0)); allowed_assets={"ACTION","ETF"} if not cfg.get("gold_in_virtual_pea_book",False) else {"ACTION","ETF","GOLD"}
    current=decisions[decisions["asset_class"].astype(str).isin(allowed_assets)].copy(); current["key"]=current["asset_class"].astype(str)+"|"+current["horizon"].astype(str)+"|"+current["isin"].astype(str)
    prev_map=dict(zip(previous.get("key",pd.Series(dtype=str)),previous.get("decision",pd.Series(dtype=str)))) if not previous.empty else {}
    new_signals=[]
    for _,row in current.iterrows():
        decision=str(row.get("decision","") or ""); key=row["key"]; score=_num(row.get("score"))
        if decision not in buy_set or prev_map.get(key) in buy_set: continue
        if score is None or score < min_buy_score: continue
        isin=str(row.get("isin","") or ""); p=prices.get(isin,{}).get("price")
        if p is None: continue
        sid=f"{today}|{key}"; new_signals.append({"signal_id":sid,"signal_date":today,"asset_class":row.get("asset_class"),"horizon":row.get("horizon"),"isin":isin,"name":row.get("name"),"sector":row.get("sector"),"entry_price":p,"entry_score":score,"entry_decision":decision})
    if new_signals: signals=pd.concat([signals,pd.DataFrame(new_signals)],ignore_index=True).drop_duplicates("signal_id",keep="first")

    mark_rows=[]
    for _,s in signals.iterrows():
        p=prices.get(str(s.get("isin","") or ""),{}).get("price"); entry=_num(s.get("entry_price"))
        if p is None or entry is None or entry<=0: continue
        pnl=(p/entry-1.0)*100.0; mark_rows.append({"date":today,"signal_id":s.get("signal_id"),"asset_class":s.get("asset_class"),"horizon":s.get("horizon"),"isin":s.get("isin"),"name":s.get("name"),"sector":s.get("sector"),"entry_price":entry,"current_price":p,"performance_pct":pnl,"positive":pnl>0})
    if mark_rows:
        marks=pd.concat([marks,pd.DataFrame(mark_rows)],ignore_index=True); marks=marks.drop_duplicates(["date","signal_id"],keep="last")

    cash=float(pd.to_numeric(nav.get("cash_eur",pd.Series(dtype=float)),errors="coerce").dropna().iloc[-1]) if not nav.empty and pd.to_numeric(nav.get("cash_eur"),errors="coerce").notna().any() else float(cfg["initial_capital_eur"])
    if positions.empty: positions=pd.DataFrame(columns=["position_id","open_date","asset_class","horizon","isin","name","sector","entry_price","quantity","entry_score","stop_pct","status"])
    current_map=current.set_index("key",drop=False).to_dict("index") if not current.empty else {}
    fee=float(cfg["transaction_cost_per_side_pct"])/100.0
    for idx,pos in positions[positions.get("status",pd.Series(index=positions.index,dtype=str)).astype(str)=="OPEN"].iterrows():
        isin=str(pos.get("isin","") or ""); p=prices.get(isin,{}).get("price"); entry=_num(pos.get("entry_price")); qty=_num(pos.get("quantity")); stop=_num(pos.get("stop_pct"))
        if p is None or entry is None or qty is None: continue
        key=f"{pos.get('asset_class')}|{pos.get('horizon')}|{isin}"; decision=str(current_map.get(key,{}).get("decision",""))
        stop_hit=stop is not None and p<=entry*(1.0-stop/100.0); decision_exit=decision in set(cfg["exit_on_decisions"])
        if stop_hit or decision_exit:
            proceeds=qty*p*(1.0-fee); cash+=proceeds; positions.at[idx,"status"]="CLOSED"; positions.at[idx,"close_date"]=today; positions.at[idx,"close_price"]=p; positions.at[idx,"exit_reason"]="STOP" if stop_hit else f"DECISION_{decision}"; tx=pd.concat([tx,pd.DataFrame([{"date":today,"type":"SELL","position_id":pos.get("position_id"),"isin":isin,"price":p,"quantity":qty,"gross_eur":qty*p,"cost_eur":qty*p*fee,"reason":positions.at[idx,"exit_reason"]}])],ignore_index=True)

    open_positions=positions[positions["status"].astype(str)=="OPEN"].copy(); exposure=sum((_num(r.get("quantity")) or 0)*(prices.get(str(r.get("isin","") or ""),{}).get("price") or (_num(r.get("entry_price")) or 0)) for _,r in open_positions.iterrows()); equity=max(0.0,cash+exposure)
    sector_exposure={}
    for _,r in open_positions.iterrows():
        val=(_num(r.get("quantity")) or 0)*(prices.get(str(r.get("isin","") or ""),{}).get("price") or (_num(r.get("entry_price")) or 0)); sector_exposure[str(r.get("sector","NON CLASSE"))]=sector_exposure.get(str(r.get("sector","NON CLASSE")),0)+val
    open_keys=set(open_positions["asset_class"].astype(str)+"|"+open_positions["horizon"].astype(str)+"|"+open_positions["isin"].astype(str)) if not open_positions.empty else set()
    for s in new_signals:
        key=f"{s['asset_class']}|{s['horizon']}|{s['isin']}"
        if key in open_keys: continue
        p=prices.get(str(s["isin"]),{}).get("price"); score=_num(s.get("entry_score")); stop=float(cfg["stops_pct"].get(str(s["horizon"]),12.0))
        if p is None or p<=0 or score is None or score<min_buy_score or equity<=0: continue
        max_pos=equity*float(cfg["max_position_pct"])/100.0; risk_pos=equity*(float(cfg["risk_budget_per_position_pct"])/100.0)/(stop/100.0); exposure_room=max(0.0,equity*float(cfg["max_total_exposure_pct"])/100.0-exposure); sector=s.get("sector") or prices.get(str(s["isin"]),{}).get("sector") or "NON CLASSE"; sector_room=max(0.0,equity*float(cfg["max_sector_exposure_pct"])/100.0-sector_exposure.get(str(sector),0.0)); cash_room=max(0.0,cash-equity*float(cfg["cash_buffer_min_pct"])/100.0); conviction=min(1.0,max(0.70,score/100.0)); value=min(max_pos,risk_pos,exposure_room,sector_room,cash_room)*conviction
        if value<=0: continue
        qty=value/p; cost=value*(1.0+fee)
        if cost>cash: continue
        cash-=cost; exposure+=value; sector_exposure[str(sector)]=sector_exposure.get(str(sector),0)+value; pid=f"{today}|{key}"; positions=pd.concat([positions,pd.DataFrame([{"position_id":pid,"open_date":today,"asset_class":s["asset_class"],"horizon":s["horizon"],"isin":s["isin"],"name":s["name"],"sector":sector,"entry_price":p,"quantity":qty,"entry_score":score,"stop_pct":stop,"status":"OPEN"}])],ignore_index=True); tx=pd.concat([tx,pd.DataFrame([{"date":today,"type":"BUY","position_id":pid,"isin":s["isin"],"price":p,"quantity":qty,"gross_eur":value,"cost_eur":value*fee,"reason":"COMMITTEE_BUY"}])],ignore_index=True); open_keys.add(key)

    open_positions=positions[positions["status"].astype(str)=="OPEN"].copy(); market_value=0.0
    for _,r in open_positions.iterrows(): market_value+=(_num(r.get("quantity")) or 0)*(prices.get(str(r.get("isin","") or ""),{}).get("price") or (_num(r.get("entry_price")) or 0))
    current_nav=cash+market_value; cumulative=(current_nav/float(cfg["initial_capital_eur"])-1.0)*100.0; nav_row={"date":today,"nav_eur":current_nav,"cash_eur":cash,"market_value_eur":market_value,"exposure_pct":market_value/current_nav*100.0 if current_nav else 0.0,"cumulative_performance_pct":cumulative,"open_positions":int(len(open_positions))}; nav=pd.concat([nav,pd.DataFrame([nav_row])],ignore_index=True).drop_duplicates("date",keep="last")
    latest_marks=marks.sort_values("date").groupby("signal_id",as_index=False).tail(1) if not marks.empty else pd.DataFrame(); signal_summary=signals.merge(latest_marks[[c for c in ["signal_id","current_price","performance_pct","positive"] if c in latest_marks.columns]],on="signal_id",how="left") if not signals.empty else signals
    previous=current[["key","decision"]].copy()
    for df,path in ((signals,state/"signals.csv"),(marks,state/"daily_marks.csv"),(positions,state/"positions.csv"),(tx,state/"transactions.csv"),(nav,state/"nav.csv"),(previous,state/"last_decisions.csv")): _save(df,path)
    assumptions=pd.DataFrame([{"parameter":k,"value":json.dumps(v,ensure_ascii=False) if isinstance(v,(dict,list)) else v} for k,v in cfg.items()])
    dashboard=pd.DataFrame([{"as_of":today,"initial_capital_eur":cfg["initial_capital_eur"],"nav_eur":current_nav,"cumulative_performance_pct":cumulative,"cash_eur":cash,"exposure_pct":nav_row["exposure_pct"],"open_positions":len(open_positions),"buy_signals_total":len(signals),"buy_signals_positive_now":int(pd.to_numeric(signal_summary.get("positive",pd.Series(dtype=float)),errors="coerce").fillna(0).sum()) if not signal_summary.empty else 0,"minimum_buy_score":min_buy_score}])
    xlsx=outdir/"COMMITTEE_BUY_PERFORMANCE.xlsx"
    with pd.ExcelWriter(xlsx,engine="openpyxl") as writer:
        dashboard.to_excel(writer,sheet_name="Dashboard",index=False); nav.to_excel(writer,sheet_name="NAV_Quotidienne",index=False); open_positions.to_excel(writer,sheet_name="Positions_Ouvertes",index=False); signals.to_excel(writer,sheet_name="Signaux_BUY",index=False); signal_summary.to_excel(writer,sheet_name="Performance_Signaux",index=False); marks.to_excel(writer,sheet_name="Suivi_Journalier",index=False); tx.to_excel(writer,sheet_name="Transactions_Virtuelles",index=False); assumptions.to_excel(writer,sheet_name="Money_Management",index=False)
    _style_xlsx(xlsx)
    return {"status":"SUCCESS","as_of":today,"xlsx":str(xlsx.relative_to(root)),"nav_eur":round(current_nav,2),"cumulative_performance_pct":round(cumulative,4),"open_positions":int(len(open_positions)),"signals":int(len(signals)),"minimum_buy_score":min_buy_score,"live_orders_enabled":False}
