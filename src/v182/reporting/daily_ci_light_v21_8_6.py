from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import json
import math

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from v182.reporting import daily_ci_light_v21_8_2 as base

ROOT = base.ROOT
VERSION = "DAILY_CI_LIGHT_V21_8_6_TRADINGVIEW"
TRADINGVIEW_ACCEPTED = {"BUY", "STRONG_BUY"}
MAX_TRADINGVIEW_AGE_HOURS = 48.0


def _technical_context(root: Path) -> dict[str, dict]:
    payload = base._load_json(root / "state/provenance/source_cache/TRADINGVIEW_TECHNICAL_V1.json")
    return dict(payload.get("entries") or {})


def _tv_row(isin: str, entries: dict[str, dict]) -> dict:
    entry = dict(entries.get(isin) or {})
    fields = dict(entry.get("fields") or {})
    age = base._age_hours(entry.get("fetched_at_utc"))
    daily = base._norm(fields.get("tradingview_daily_signal"))
    weekly = base._norm(fields.get("tradingview_weekly_signal"))
    monthly = base._norm(fields.get("tradingview_monthly_signal"))
    complete = all(signal in TRADINGVIEW_ACCEPTED for signal in (daily, weekly, monthly))
    fresh = age <= MAX_TRADINGVIEW_AGE_HOURS
    return {
        "tradingview_daily": daily or None,
        "tradingview_weekly": weekly or None,
        "tradingview_monthly": monthly or None,
        "tradingview_all_buy": bool(complete),
        "tradingview_fresh": bool(fresh),
        "tradingview_age_hours": None if not math.isfinite(age) else round(age, 2),
        "tradingview_url": entry.get("source_url") or None,
        "tradingview_symbol": entry.get("symbol") or None,
    }


def _build_actions(root: Path, decisions: pd.DataFrame, tv: dict[str, dict]) -> tuple[list[dict], int]:
    action_decisions, _ = base._decision_index(decisions)
    cache = base._action_cache(root)
    rows: list[dict] = []
    complete_boursorama = 0
    for isin, entry in cache.items():
        fields = dict(entry.get("fields") or {})
        analysts = base._num(fields.get("boursorama_n_analysts"))
        consensus = base._norm(fields.get("boursorama_consensus"))
        upside = base._num(fields.get("boursorama_target_upside_pct"))
        if analysts is not None and consensus and upside is not None:
            complete_boursorama += 1
        tech = _tv_row(isin, tv)
        eligible = (
            analysts is not None and analysts > base.MIN_ACTION_ANALYSTS_EXCLUSIVE
            and consensus in base.BOURSORAMA_ACTION_ACCEPTED
            and upside is not None and upside > base.MIN_ACTION_UPSIDE_EXCLUSIVE
            and tech["tradingview_all_buy"] and tech["tradingview_fresh"]
        )
        if not eligible:
            continue
        model = action_decisions.get(isin, {})
        rows.append({
            "asset_class": "ACTION", "isin": isin, "name": model.get("name"),
            "decision_ct": model.get("decision_ct"), "score_ct": model.get("score_ct"),
            "decision_tct": model.get("decision_tct"), "score_tct": model.get("score_tct"),
            "boursorama_n_analysts": analysts, "boursorama_recommendation": consensus,
            "boursorama_target_upside_pct": upside, "morningstar_rating": None,
            **tech,
            "boursorama_url": entry.get("consensus_url") or entry.get("key_figures_url"),
            "selection_rule": "ANALYSTS>10 + BOURSORAMA_ACHETER/RENFORCER + UPSIDE>20% + TRADINGVIEW_DAY/WEEK/MONTH_BUY",
        })
    return rows, complete_boursorama


def _build_etfs(root: Path, decisions: pd.DataFrame, tv: dict[str, dict]) -> tuple[list[dict], int]:
    _, etf_decisions = base._decision_index(decisions)
    master = base._etf_master(root)
    cache = base._etf_cache(root)
    rows: list[dict] = []
    proven = 0
    if master.empty or "isin" not in master.columns:
        return rows, proven
    for row in master.to_dict("records"):
        isin = str(row.get("isin") or "").strip()
        if not isin:
            continue
        cache_entry = dict(cache.get(isin) or {})
        fields = dict(cache_entry.get("fields") or {})
        rating = base._stars(fields.get("boursorama_etf_morningstar_rating"))
        proof_valid = fields.get("boursorama_morningstar_rating_proof_valid") is True
        proof_url = str(fields.get("boursorama_morningstar_rating_source_url") or cache_entry.get("course_url") or "").strip()
        if rating is not None and proof_valid and "boursorama.com" in proof_url.lower():
            proven += 1
        else:
            continue
        tech = _tv_row(isin, tv)
        eligible = rating > 3.0 and tech["tradingview_all_buy"] and tech["tradingview_fresh"]
        if not eligible:
            continue
        model = etf_decisions.get(isin, {})
        rows.append({
            "asset_class": "ETF", "isin": isin, "name": model.get("name") or row.get("name"),
            "decision_ct": model.get("decision_ct"), "score_ct": model.get("score_ct"),
            "decision_tct": None, "score_tct": None,
            "boursorama_n_analysts": None, "boursorama_recommendation": None,
            "boursorama_target_upside_pct": None, "morningstar_rating": rating,
            **tech, "boursorama_url": proof_url,
            "selection_rule": "BOURSORAMA_PROVEN_MORNINGSTAR>3_STARS + TRADINGVIEW_DAY/WEEK/MONTH_BUY",
        })
    return rows, proven


def _write_excel(path: Path, frame: pd.DataFrame) -> None:
    if not path.exists():
        return
    wb = load_workbook(path)
    if "CI_LIGHT" in wb.sheetnames:
        del wb["CI_LIGHT"]
    ws = wb.create_sheet("CI_LIGHT")
    columns = list(frame.columns) if not frame.empty else [
        "asset_class","isin","name","decision_ct","score_ct","decision_tct","score_tct",
        "boursorama_n_analysts","boursorama_recommendation","boursorama_target_upside_pct","morningstar_rating",
        "tradingview_daily","tradingview_weekly","tradingview_monthly","tradingview_symbol",
        "boursorama_url","tradingview_url","tradingview_age_hours","selection_rule"
    ]
    for idx, col in enumerate(columns, 1):
        c = ws.cell(1, idx, col); c.font = Font(bold=True); c.alignment = Alignment(wrap_text=True, vertical="top")
    for ridx, record in enumerate(frame.to_dict("records"), 2):
        for cidx, col in enumerate(columns, 1):
            value = record.get(col)
            try:
                if pd.isna(value): value = None
            except Exception:
                pass
            cell = ws.cell(ridx, cidx, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col in {"boursorama_url", "tradingview_url"} and value:
                cell.hyperlink = str(value); cell.style = "Hyperlink"
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def _append_android(path: Path, frame: pd.DataFrame) -> None:
    lines = ["", "---", "", "## CI LIGHT — Boursorama + TradingView", "",
             "Actions : >10 analystes, Boursorama ACHETER/RENFORCER, potentiel >20 %, TradingView BUY/STRONG_BUY jour + semaine + mois.",
             "ETF : Morningstar Boursorama >3★ avec preuve Boursorama, TradingView BUY/STRONG_BUY jour + semaine + mois.", ""]
    if frame.empty:
        lines.append("Aucun instrument ne satisfait simultanément tous les critères vérifiés.")
    else:
        lines += ["| Type | Instrument | Boursorama | Potentiel/★ | TV J | TV H | TV M |", "|---|---|---|---:|---|---|---|"]
        for row in frame.to_dict("records"):
            if row.get("asset_class") == "ACTION":
                b = f"{int(row['boursorama_n_analysts'])} analystes · {row['boursorama_recommendation']}"; metric = f"{row['boursorama_target_upside_pct']:.1f}%"
            else:
                b = "Morningstar Boursorama"; metric = f"{row['morningstar_rating']:.1f}★"
            lines.append(f"| {row.get('asset_class')} | {row.get('name') or row.get('isin')} | {b} | {metric} | {row.get('tradingview_daily')} | {row.get('tradingview_weekly')} | {row.get('tradingview_monthly')} |")
        lines += ["", "Les liens Boursorama et TradingView figurent dans l'onglet Excel `CI_LIGHT` et le CSV."]
    with path.open("a", encoding="utf-8") as f: f.write("\n".join(lines) + "\n")


def run(root: Path = ROOT) -> dict:
    decisions = base._read_csv(root / "outputs/daily_tct_ct/DAILY_TCT_CT_V21_8.csv")
    if decisions.empty:
        raise RuntimeError("CI_LIGHT_DECISIONS_MISSING")
    tv = _technical_context(root)
    actions, action_complete = _build_actions(root, decisions, tv)
    etfs, etf_proven = _build_etfs(root, decisions, tv)
    frame = pd.DataFrame(actions + etfs)
    if not frame.empty:
        frame["_asset"] = frame["asset_class"].map({"ACTION":0,"ETF":1}).fillna(9)
        frame["_metric"] = pd.to_numeric(frame["boursorama_target_upside_pct"], errors="coerce").fillna(pd.to_numeric(frame["morningstar_rating"], errors="coerce"))
        frame = frame.sort_values(["_asset","_metric","score_ct"], ascending=[True,False,False]).drop(columns=["_asset","_metric"])

    outdir=root/"outputs/daily_tct_ct"; auditdir=root/"outputs/audit"; committee=root/"outputs/committee_master"; mobile=root/"outputs/mobile"
    for d in (outdir,auditdir,committee,mobile): d.mkdir(parents=True,exist_ok=True)
    csv_path=outdir/"CI_LIGHT_V21_8_6.csv"; frame.to_csv(csv_path,sep=";",index=False,encoding="utf-8-sig")
    excel=committee/"CI_REFERENTIEL_PONDERE.xlsx"; _write_excel(excel,frame)
    android=mobile/"ANDROID_CI_CONTROL_CENTER.md"
    if android.exists(): _append_android(android,frame)

    complete=positive=0
    for isin in tv:
        r=_tv_row(isin,tv)
        if all(r[k] is not None for k in ("tradingview_daily","tradingview_weekly","tradingview_monthly")): complete+=1
        if r["tradingview_all_buy"] and r["tradingview_fresh"]: positive+=1
    payload={
        "status":"SUCCESS","version":VERSION,"generated_at_utc":datetime.now(timezone.utc).isoformat(),
        "selection_is_additional_ci_light_only":True,"full_ci_preserved":True,"full_ci_weighted_criteria_preserved":True,
        "technical_provider":"TradingView","investing_active":False,
        "action_rule":{"boursorama_n_analysts":">10","boursorama_recommendation":sorted(base.BOURSORAMA_ACTION_ACCEPTED),"boursorama_target_upside_pct":">20","tradingview_daily":sorted(TRADINGVIEW_ACCEPTED),"tradingview_weekly":sorted(TRADINGVIEW_ACCEPTED),"tradingview_monthly":sorted(TRADINGVIEW_ACCEPTED)},
        "etf_rule":{"analyst_recommendation_required":False,"boursorama_morningstar_rating":">3 stars with positive Boursorama proof","tradingview_daily":sorted(TRADINGVIEW_ACCEPTED),"tradingview_weekly":sorted(TRADINGVIEW_ACCEPTED),"tradingview_monthly":sorted(TRADINGVIEW_ACCEPTED)},
        "max_tradingview_age_hours":MAX_TRADINGVIEW_AGE_HOURS,"action_boursorama_complete_context":int(action_complete),"etf_boursorama_morningstar_positive_proof":int(etf_proven),
        "tradingview_cache_entries":len(tv),"tradingview_three_horizon_complete_entries":complete,"tradingview_three_horizon_all_buy_fresh_entries":positive,
        "selected_rows":len(frame),"selected_actions":int((frame.get("asset_class",pd.Series(dtype=str))=="ACTION").sum()),"selected_etfs":int((frame.get("asset_class",pd.Series(dtype=str))=="ETF").sum()),
        "csv_output":str(csv_path.relative_to(root)),"excel_output":str(excel.relative_to(root)),"excel_sheet":"CI_LIGHT","android_output":str(android.relative_to(root)) if android.exists() else None,
        "external_collection_calls":0,"score_or_decision_mutation":False,"weights_changed":False,"selection_thresholds_changed":False,"real_orders_enabled":False
    }
    (auditdir/"DAILY_CI_LIGHT_V21_8_6.json").write_text(json.dumps(payload,ensure_ascii=False,indent=2,default=str),encoding="utf-8")
    return payload


if __name__ == "__main__": print(json.dumps(run(),ensure_ascii=False,indent=2,default=str))
