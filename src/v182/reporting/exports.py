from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from v182.io.frames import is_missing

HEADER_FILL="17365D"

def _coverage(frame: pd.DataFrame) -> pd.DataFrame:
    rows=[]
    for col in frame.columns:
        observed=int((~frame[col].apply(is_missing)).sum())
        rows.append({"column":col,"observed":observed,"missing":len(frame)-observed,"coverage_pct":round(observed/len(frame)*100,1) if len(frame) else 0})
    return pd.DataFrame(rows).sort_values(["coverage_pct","column"])

def _format(path: Path) -> None:
    from openpyxl import load_workbook
    wb=load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes="A2"
        for cell in ws[1]:
            cell.font=Font(bold=True,color="FFFFFF")
            cell.fill=PatternFill("solid",fgColor=HEADER_FILL)
        for i,col in enumerate(ws.columns,1):
            max_len=min(max((len(str(c.value)) if c.value is not None else 0) for c in col),60)
            ws.column_dimensions[get_column_letter(i)].width=max(12,max_len+2)
    wb.save(path)

def export_master_excel(frame: pd.DataFrame, path: str | Path, title: str) -> Path:
    path=Path(path); path.parent.mkdir(parents=True,exist_ok=True)
    with pd.ExcelWriter(path,engine="openpyxl") as writer:
        pd.DataFrame({"indicator":["title","rows","columns"],"value":[title,len(frame),len(frame.columns)]}).to_excel(writer,sheet_name="Dashboard",index=False)
        frame.to_excel(writer,sheet_name="Referentiel",index=False)
        _coverage(frame).to_excel(writer,sheet_name="Couverture_colonnes",index=False)
    _format(path); return path

def export_run_report(before: dict, after: dict, quality_checks: list[dict], path: str | Path) -> Path:
    path=Path(path); path.parent.mkdir(parents=True,exist_ok=True)
    summary=[]
    for u in ("ACTION","ETF"):
        summary.append({"universe":u,"coverage_before_pct":before[u]["coverage_pct"],"coverage_after_pct":after[u]["coverage_pct"],"gain_points":round(after[u]["coverage_pct"]-before[u]["coverage_pct"],2)})
    with pd.ExcelWriter(path,engine="openpyxl") as writer:
        pd.DataFrame(summary).to_excel(writer,sheet_name="Couverture",index=False)
        pd.DataFrame(quality_checks).to_excel(writer,sheet_name="Quality_Gates",index=False)
    _format(path); return path
