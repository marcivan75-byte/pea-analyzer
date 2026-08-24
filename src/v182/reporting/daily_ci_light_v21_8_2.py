from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import json
import math
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


ROOT = Path(__file__).resolve().parents[3]
VERSION = "DAILY_CI_LIGHT_V21_8_2"
INVESTING_ACCEPTED = {"BUY", "STRONG_BUY"}
BOURSORAMA_ACTION_ACCEPTED = {"BUY", "STRONG_BUY", "ACHETER", "RENFORCER"}
MIN_ACTION_ANALYSTS_EXCLUSIVE = 10
MIN_ACTION_UPSIDE_EXCLUSIVE = 20.0
MIN_ETF_MORNINGSTAR_EXCLUSIVE = 3.0
MAX_INVESTING_AGE_HOURS = 48.0


def _read_csv(path: Path) -> pd.DataFrame:
    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame()
    return pd.read_csv(path, sep=";", encoding="utf-8-sig", low_memory=False)


def _load_json(path: Path) -> dict:
    if not path.exists() or path.stat().st_size == 0:
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError, json.JSONDecodeError):
        return {}


def _norm(value: object) -> str:
    return "_".join(str(value or "").strip().upper().replace("-", " ").split())


def _num(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace("\u202f", " ").replace("\xa0", " ")
    if not text or text.lower() in {"nan", "none", "null"}:
        return None
    text = text.replace("%", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".") if text.rfind(",") > text.rfind(".") else text.replace(",", "")
    else:
        text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def _stars(value: object) -> float | None:
    direct = _num(value)
    if direct is not None and 0.0 <= direct <= 5.0:
        return direct
    match = re.search(r"([0-5](?:[.,]\d+)?)", str(value or ""))
    if not match:
        return None
    return _num(match.group(1))


def _age_hours(value: object) -> float:
    text = str(value or "").strip()
    if not text:
        return math.inf
    try:
        stamp = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return math.inf
    if stamp.tzinfo is None:
        stamp = stamp.replace(tzinfo=timezone.utc)
    return max(0.0, (datetime.now(timezone.utc) - stamp.astimezone(timezone.utc)).total_seconds() / 3600.0)


def _investing_context(root: Path) -> tuple[dict[str, dict], dict[str, dict]]:
    cache = _load_json(root / "state" / "provenance" / "source_cache" / "INVESTING_TECHNICAL_V1.json")
    mappings = _load_json(root / "state" / "provenance" / "source_cache" / "INVESTING_URL_MAP_V1.json")
    entries = dict(cache.get("entries") or {})
    url_map = dict(mappings.get("entries") or {})
    return entries, url_map


def _investing_row(isin: str, entries: dict[str, dict], url_map: dict[str, dict]) -> dict:
    entry = dict(entries.get(isin) or {})
    fields = dict(entry.get("fields") or {})
    age = _age_hours(entry.get("fetched_at_utc"))
    daily = _norm(fields.get("investing_daily_signal"))
    weekly = _norm(fields.get("investing_weekly_signal"))
    monthly = _norm(fields.get("investing_monthly_signal"))
    complete = all(signal in INVESTING_ACCEPTED for signal in (daily, weekly, monthly))
    fresh = age <= MAX_INVESTING_AGE_HOURS
    mapping = dict(url_map.get(isin) or {})
    base_url = str(mapping.get("base_url") or "").strip()
    technical_url = str(entry.get("source_url") or "").strip()
    if not technical_url and base_url:
        technical_url = base_url.rstrip("/") + "-technical"
    return {
        "investing_daily": daily or None,
        "investing_weekly": weekly or None,
        "investing_monthly": monthly or None,
        "investing_all_buy": bool(complete),
        "investing_fresh": bool(fresh),
        "investing_age_hours": None if not math.isfinite(age) else round(age, 2),
        "investing_url": base_url or technical_url or None,
        "investing_technical_url": technical_url or None,
    }


def _action_cache(root: Path) -> dict[str, dict]:
    payload = _load_json(root / "state" / "provenance" / "source_cache" / "BOURSORAMA_SELECTED_V1.json")
    return dict(payload.get("entries") or {})


def _etf_cache(root: Path) -> dict[str, dict]:
    payload = _load_json(root / "state" / "provenance" / "source_cache" / "BOURSORAMA_SELECTED_ETF_V1.json")
    return dict(payload.get("entries") or {})


def _etf_master(root: Path) -> pd.DataFrame:
    parquet = root / "state" / "provenance" / "daily_fast_master_v1" / "etf.parquet"
    csv = root / "outputs" / "V18.2_PEA_ETF_MASTER_ENRICHED.csv"
    if parquet.exists():
        try:
            return pd.read_parquet(parquet)
        except Exception:
            pass
    return _read_csv(csv)


def _decision_index(decisions: pd.DataFrame) -> tuple[dict[str, dict], dict[str, dict]]:
    if decisions.empty:
        return {}, {}
    work = decisions.copy()
    work["asset_class"] = work.get("asset_class", "").astype(str).str.upper()
    work["horizon"] = work.get("horizon", "").astype(str).str.upper()
    action_ct = work[(work["asset_class"] == "ACTION") & (work["horizon"] == "CT")]
    action_tct = work[(work["asset_class"] == "ACTION") & (work["horizon"] == "TCT")]
    etf_ct = work[(work["asset_class"] == "ETF") & (work["horizon"] == "CT")]
    action_tct_map = {str(r.get("isin") or ""): r for r in action_tct.to_dict("records")}
    actions: dict[str, dict] = {}
    for row in action_ct.to_dict("records"):
        isin = str(row.get("isin") or "")
        tct = action_tct_map.get(isin, {})
        actions[isin] = {
            "isin": isin,
            "name": row.get("name") or tct.get("name"),
            "decision_ct": row.get("decision"),
            "score_ct": _num(row.get("score")),
            "decision_tct": tct.get("decision"),
            "score_tct": _num(tct.get("score")),
        }
    etfs = {
        str(row.get("isin") or ""): {
            "isin": str(row.get("isin") or ""),
            "name": row.get("name"),
            "decision_ct": row.get("decision"),
            "score_ct": _num(row.get("score")),
        }
        for row in etf_ct.to_dict("records")
    }
    return actions, etfs


def _build_actions(root: Path, decisions: pd.DataFrame, investing: dict[str, dict], url_map: dict[str, dict]) -> tuple[list[dict], int]:
    action_decisions, _ = _decision_index(decisions)
    cache = _action_cache(root)
    rows: list[dict] = []
    complete_boursorama = 0
    for isin, entry in cache.items():
        fields = dict(entry.get("fields") or {})
        analysts = _num(fields.get("boursorama_n_analysts"))
        consensus = _norm(fields.get("boursorama_consensus"))
        upside = _num(fields.get("boursorama_target_upside_pct"))
        if analysts is not None and consensus and upside is not None:
            complete_boursorama += 1
        inv = _investing_row(isin, investing, url_map)
        eligible = (
            analysts is not None
            and analysts > MIN_ACTION_ANALYSTS_EXCLUSIVE
            and consensus in BOURSORAMA_ACTION_ACCEPTED
            and upside is not None
            and upside > MIN_ACTION_UPSIDE_EXCLUSIVE
            and inv["investing_all_buy"]
            and inv["investing_fresh"]
        )
        if not eligible:
            continue
        model = action_decisions.get(isin, {})
        rows.append({
            "asset_class": "ACTION",
            "isin": isin,
            "name": model.get("name"),
            "decision_ct": model.get("decision_ct"),
            "score_ct": model.get("score_ct"),
            "decision_tct": model.get("decision_tct"),
            "score_tct": model.get("score_tct"),
            "boursorama_n_analysts": analysts,
            "boursorama_recommendation": consensus,
            "boursorama_target_upside_pct": upside,
            "morningstar_rating": None,
            "investing_daily": inv["investing_daily"],
            "investing_weekly": inv["investing_weekly"],
            "investing_monthly": inv["investing_monthly"],
            "boursorama_url": entry.get("consensus_url") or entry.get("key_figures_url"),
            "investing_url": inv["investing_url"],
            "investing_technical_url": inv["investing_technical_url"],
            "investing_age_hours": inv["investing_age_hours"],
            "selection_rule": "ANALYSTS>10 + BOURSORAMA_BUY/STRONG_BUY + UPSIDE>20% + INVESTING_DAY/WEEK/MONTH_BUY",
        })
    return rows, complete_boursorama


def _boursorama_evidence(row: dict, cache_entry: dict) -> tuple[bool, str | None]:
    urls = [
        cache_entry.get("composition_url"),
        cache_entry.get("risk_url"),
        row.get("source_url"),
        row.get("ticker_source_url_final"),
    ]
    sources = [row.get("source"), row.get("source_name")]
    evidence = any("boursorama.com" in str(value or "").lower() for value in urls) or any(
        "boursorama" in str(value or "").lower() for value in sources
    )
    url = next((str(value) for value in urls if "boursorama.com" in str(value or "").lower()), None)
    return bool(evidence), url


def _build_etfs(root: Path, decisions: pd.DataFrame, investing: dict[str, dict], url_map: dict[str, dict]) -> tuple[list[dict], int]:
    _, etf_decisions = _decision_index(decisions)
    master = _etf_master(root)
    cache = _etf_cache(root)
    rows: list[dict] = []
    boursorama_rated = 0
    if master.empty or "isin" not in master.columns:
        return rows, boursorama_rated
    for row in master.to_dict("records"):
        isin = str(row.get("isin") or "")
        if not isin:
            continue
        cache_entry = dict(cache.get(isin) or {})
        cache_fields = dict(cache_entry.get("fields") or {})
        rating = _stars(cache_fields.get("boursorama_etf_morningstar_rating"))
        if rating is None:
            rating = _stars(row.get("morningstar_rating"))
        has_boursorama, boursorama_url = _boursorama_evidence(row, cache_entry)
        if rating is not None and has_boursorama:
            boursorama_rated += 1
        inv = _investing_row(isin, investing, url_map)
        eligible = (
            rating is not None
            and rating > MIN_ETF_MORNINGSTAR_EXCLUSIVE
            and has_boursorama
            and inv["investing_all_buy"]
            and inv["investing_fresh"]
        )
        if not eligible:
            continue
        model = etf_decisions.get(isin, {})
        rows.append({
            "asset_class": "ETF",
            "isin": isin,
            "name": model.get("name") or row.get("name"),
            "decision_ct": model.get("decision_ct"),
            "score_ct": model.get("score_ct"),
            "decision_tct": None,
            "score_tct": None,
            "boursorama_n_analysts": None,
            "boursorama_recommendation": None,
            "boursorama_target_upside_pct": None,
            "morningstar_rating": rating,
            "investing_daily": inv["investing_daily"],
            "investing_weekly": inv["investing_weekly"],
            "investing_monthly": inv["investing_monthly"],
            "boursorama_url": boursorama_url,
            "investing_url": inv["investing_url"],
            "investing_technical_url": inv["investing_technical_url"],
            "investing_age_hours": inv["investing_age_hours"],
            "selection_rule": "BOURSORAMA_MORNINGSTAR>3_STARS + INVESTING_DAY/WEEK/MONTH_BUY",
        })
    return rows, boursorama_rated


def _write_excel_sheet(path: Path, frame: pd.DataFrame) -> None:
    if not path.exists():
        return
    wb = load_workbook(path)
    if "CI_LIGHT" in wb.sheetnames:
        del wb["CI_LIGHT"]
    ws = wb.create_sheet("CI_LIGHT")
    columns = list(frame.columns)
    if not columns:
        columns = [
            "asset_class", "isin", "name", "decision_ct", "score_ct", "decision_tct", "score_tct",
            "boursorama_n_analysts", "boursorama_recommendation", "boursorama_target_upside_pct",
            "morningstar_rating", "investing_daily", "investing_weekly", "investing_monthly",
            "boursorama_url", "investing_url", "investing_technical_url", "investing_age_hours", "selection_rule",
        ]
    for col_idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row_idx, record in enumerate(frame.to_dict("records"), start=2):
        for col_idx, column in enumerate(columns, start=1):
            value = record.get(column)
            if pd.isna(value) if not isinstance(value, (list, dict, tuple, set)) else False:
                value = None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if column in {"boursorama_url", "investing_url", "investing_technical_url"} and value:
                cell.hyperlink = str(value)
                cell.style = "Hyperlink"
    widths = {
        "A": 14, "B": 16, "C": 34, "D": 18, "E": 12, "F": 18, "G": 12,
        "H": 22, "I": 28, "J": 24, "K": 20, "L": 18, "M": 18, "N": 18,
        "O": 54, "P": 54, "Q": 54, "R": 20, "S": 72,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def _append_android(path: Path, frame: pd.DataFrame) -> None:
    lines = [
        "",
        "---",
        "",
        "## CI LIGHT — filtre externe strict",
        "",
        "Actions : >10 analystes, Boursorama BUY/STRONG_BUY, potentiel >20 %, Investing BUY/STRONG_BUY jour + semaine + mois.",
        "ETF : Morningstar Boursorama >3 étoiles, Investing BUY/STRONG_BUY jour + semaine + mois.",
        "",
    ]
    if frame.empty:
        lines.append("Aucun instrument ne satisfait simultanément tous les critères avec données vérifiées et suffisamment fraîches.")
    else:
        lines.extend([
            "| Type | Instrument | Boursorama | Potentiel/★ | Investing J | H | M |",
            "|---|---|---|---:|---|---|---|",
        ])
        for row in frame.to_dict("records"):
            if row.get("asset_class") == "ACTION":
                b = f"{int(row['boursorama_n_analysts'])} analystes · {row['boursorama_recommendation']}"
                metric = f"{row['boursorama_target_upside_pct']:.1f}%"
            else:
                b = "Morningstar"
                metric = f"{row['morningstar_rating']:.1f}★"
            lines.append(
                f"| {row.get('asset_class')} | {row.get('name') or row.get('isin')} | {b} | {metric} | "
                f"{row.get('investing_daily')} | {row.get('investing_weekly')} | {row.get('investing_monthly')} |"
            )
        lines.extend(["", "Les URLs Boursorama et Investing figurent dans l'onglet Excel `CI_LIGHT` et le CSV CI LIGHT."])
    with path.open("a", encoding="utf-8") as handle:
        handle.write("\n".join(lines) + "\n")


def run(root: Path = ROOT) -> dict:
    decisions_path = root / "outputs" / "daily_tct_ct" / "DAILY_TCT_CT_V21_8.csv"
    decisions = _read_csv(decisions_path)
    if decisions.empty:
        raise RuntimeError("CI_LIGHT_DECISIONS_MISSING")

    investing, url_map = _investing_context(root)
    action_rows, action_complete_boursorama = _build_actions(root, decisions, investing, url_map)
    etf_rows, etf_boursorama_rated = _build_etfs(root, decisions, investing, url_map)
    frame = pd.DataFrame(action_rows + etf_rows)
    if not frame.empty:
        frame["_sort"] = frame["asset_class"].map({"ACTION": 0, "ETF": 1}).fillna(9)
        frame["_metric"] = pd.to_numeric(frame["boursorama_target_upside_pct"], errors="coerce").fillna(
            pd.to_numeric(frame["morningstar_rating"], errors="coerce")
        )
        frame = frame.sort_values(["_sort", "_metric", "score_ct"], ascending=[True, False, False]).drop(columns=["_sort", "_metric"])

    outdir = root / "outputs" / "daily_tct_ct"
    auditdir = root / "outputs" / "audit"
    committee_dir = root / "outputs" / "committee_master"
    mobile_dir = root / "outputs" / "mobile"
    for directory in (outdir, auditdir, committee_dir, mobile_dir):
        directory.mkdir(parents=True, exist_ok=True)

    csv_path = outdir / "CI_LIGHT_V21_8_2.csv"
    frame.to_csv(csv_path, sep=";", index=False, encoding="utf-8-sig")
    excel_path = committee_dir / "CI_REFERENTIEL_PONDERE.xlsx"
    _write_excel_sheet(excel_path, frame)
    android_path = mobile_dir / "ANDROID_CI_CONTROL_CENTER.md"
    if android_path.exists():
        _append_android(android_path, frame)

    complete_investing = 0
    fresh_complete_investing = 0
    for isin in investing:
        inv = _investing_row(isin, investing, url_map)
        if all(inv[key] is not None for key in ("investing_daily", "investing_weekly", "investing_monthly")):
            complete_investing += 1
        if inv["investing_all_buy"] and inv["investing_fresh"]:
            fresh_complete_investing += 1

    payload = {
        "status": "SUCCESS",
        "version": VERSION,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "selection_is_additional_ci_light_only": True,
        "full_ci_preserved": True,
        "full_ci_weighted_criteria_preserved": True,
        "action_rule": {
            "boursorama_n_analysts": ">10",
            "boursorama_recommendation": sorted(BOURSORAMA_ACTION_ACCEPTED),
            "boursorama_target_upside_pct": ">20",
            "investing_daily": sorted(INVESTING_ACCEPTED),
            "investing_weekly": sorted(INVESTING_ACCEPTED),
            "investing_monthly": sorted(INVESTING_ACCEPTED),
        },
        "etf_rule": {
            "analyst_recommendation_required": False,
            "boursorama_morningstar_rating": ">3 stars",
            "investing_daily": sorted(INVESTING_ACCEPTED),
            "investing_weekly": sorted(INVESTING_ACCEPTED),
            "investing_monthly": sorted(INVESTING_ACCEPTED),
        },
        "max_investing_age_hours": MAX_INVESTING_AGE_HOURS,
        "action_boursorama_complete_context": int(action_complete_boursorama),
        "etf_boursorama_morningstar_evidence": int(etf_boursorama_rated),
        "investing_cache_entries": int(len(investing)),
        "investing_three_horizon_complete_entries": int(complete_investing),
        "investing_three_horizon_all_buy_fresh_entries": int(fresh_complete_investing),
        "selected_rows": int(len(frame)),
        "selected_actions": int((frame.get("asset_class", pd.Series(dtype=str)) == "ACTION").sum()),
        "selected_etfs": int((frame.get("asset_class", pd.Series(dtype=str)) == "ETF").sum()),
        "csv_output": str(csv_path.relative_to(root)),
        "excel_output": str(excel_path.relative_to(root)),
        "excel_sheet": "CI_LIGHT",
        "android_output": str(android_path.relative_to(root)) if android_path.exists() else None,
        "external_collection_calls": 0,
        "score_or_decision_mutation": False,
        "weights_changed": False,
        "selection_thresholds_changed": False,
        "real_orders_enabled": False,
    }
    (auditdir / "DAILY_CI_LIGHT_V21_8_2.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    return payload


if __name__ == "__main__":
    print(json.dumps(run(), ensure_ascii=False, indent=2, default=str))
