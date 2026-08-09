from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import json
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT=Path(__file__).resolve().parents[3]
MASTER=ROOT/'outputs/V20.4_ACTIONS_PEA_FINAL_REFERENCE_ENRICHED.csv'
OUT=ROOT/'outputs/V20.4_GITOK_ACTIONS_1829_DECISIONS.csv'
PACK=ROOT/'outputs/V20.4_GITOK_ACTIONS_1829_ABOVE_77.xlsx'
AUDIT=ROOT/'outputs/audit/V20.4_ACTIONS_1829_COMMITTEE_AUDIT.json'

FAMILIES={
 'quality':[('roe',True,1.0),('roa',True,.7),('fcf_yield',True,1.0),('marge_ebit',True,.8),('marge_nette',True,.7),('croiss_ca_3y',True,.7),('croiss_eps_3y',True,.8),('dette_ebitda',False,.8),('debt_to_equity',False,.5)],
 'value':[('per_forward',False,1.0),('per_ttm',False,.8),('pb',False,.5),('ev_ebit',False,.8),('fcf_yield',True,1.0),('per_vs_sector_pct',False,.6),('target_upside_pct',True,.6)],
 'momentum':[('perf_1m_pct',True,1.0),('perf_3m_pct',True,1.2),('perf_6m_pct',True,1.2),('perf_1y_pct',True,.7),('relative_strength',True,1.0),('macd_hist',True,.7),('rsi14',True,.3),('rvol20',True,.4)],
 'analyst':[('analyst_momentum_score',True,1.2),('consensus_score_100',True,1.0),('target_upside_pct',True,1.0),('weighted_target_revision_30d_pct',True,.8),('weighted_consensus_delta_30d',True,.8),('revision_breadth_30d',True,.7),('net_upgrades_30d',True,.6),('consensus_confidence',True,.4)],
 'risk':[('volatility_20d',False,1.0),('volatility_60d',False,.8),('max_drawdown_1y',True,1.0),('beta',False,.5),('asymmetry',True,.8),('data_trust_pct',True,.5)],
 'structure':[('market_cap',True,.8),('volume',True,.6),('coverage_pct',True,.7),('data_trust_pct',True,.7),('v182_ticker_validation_confidence_pct',True,.4)]}

def num(df,c):
 return pd.to_numeric(df[c],errors='coerce') if c in df.columns else pd.Series(np.nan,index=df.index,dtype=float)

def rank(df,c,higher):
 p=num(df,c).rank(pct=True,method='average')*100
 return (p if higher else 100-p).fillna(50).clip(0,100)

def family(df,spec):
 vals=[]; ws=[]
 for c,h,w in spec:
  if c in df.columns: vals.append(rank(df,c,h)*w); ws.append(w)
 return sum(vals)/sum(ws) if vals else pd.Series(50.0,index=df.index)

def score(df):
 f={k:family(df,v) for k,v in FAMILIES.items()}
 for k,v in f.items(): df[f'score_{k}_100']=v.round(2)
 df['score_short_term']=(.35*f['momentum']+.25*f['analyst']+.15*f['risk']+.15*f['structure']+.10*f['quality']).round(2)
 df['score_medium_term']=(.25*f['quality']+.20*f['value']+.20*f['momentum']+.20*f['analyst']+.15*f['risk']).round(2)
 df['score_long_term']=(.35*f['quality']+.25*f['value']+.15*f['structure']+.15*f['risk']+.10*f['analyst']).round(2)
 df['short_thesis_score']=(.45*(100-f['momentum'])+.30*(100-f['quality'])+.15*(100-f['analyst'])+.10*(100-f['risk'])).round(2)
 multi=.25*df.score_short_term+.35*df.score_medium_term+.40*df.score_long_term
 base=num(df,'committee_score_with_analyst_momentum').fillna(num(df,'score_brut')).fillna(50).clip(0,100)
 pct=multi.rank(pct=True,method='average')*100
 df['committee_score_1829']=(.60*base+.40*pct).round(2)
 return df

def timing(df):
 conf=num(df,'v182_ticker_validation_confidence_pct').fillna(0)/100
 qconf=num(df,'quarantine_identity_confidence').fillna(0)
 conf=np.maximum(conf,qconf)
 trust=num(df,'data_trust_pct').fillna(0)/100
 last=num(df,'last_close'); atr=num(df,'atr14'); target=num(df,'target_price').fillna(num(df,'target_mean_yf')); inval=num(df,'invalidation_level')
 reliable=(conf>=.92)&(trust>=.50)&last.notna()&(last>0)
 df['identity_method']=np.where(df.get('final_reference_origin','').astype(str).eq('QUARANTINE_506'),'QUARANTINE_API_YFINANCE','CANONICAL_TICKER_ISIN')
 df['identity_confidence']=pd.Series(conf,index=df.index).round(3)
 df['T1_entry_low']=np.where(reliable&atr.notna(),(last-.35*atr).clip(lower=0),np.nan)
 df['T1_entry_high']=np.where(reliable&atr.notna(),last+.10*atr,np.nan)
 df['T1_target']=np.where(reliable&target.notna(),target,np.nan)
 df['T1_invalidation']=np.where(reliable&inval.notna(),inval,np.where(reliable&atr.notna(),(last-1.6*atr).clip(lower=0),np.nan))
 s=df['committee_score_1829']
 df['decision']=np.select([s>77,s>=70,s>=60],['BUY_CANDIDATE','WATCH','REVIEW'],default='REJECT')
 low=pd.Series(conf,index=df.index)<.92
 df.loc[low,'decision']=np.where(s[low]>=70,'REVIEW','REJECT')
 df['execution']='RESEARCH_ONLY'
 df['T0']=np.select([df.decision.eq('BUY_CANDIDATE'),df.decision.eq('WATCH')],['PREPARE','WATCH'],default='NO_ACTION')
 df['T1_1_4w']=np.where(df.decision.eq('BUY_CANDIDATE'),'ENTRY_IF_ZONE_VALID','MONITOR')
 df['T2_1_3m']=np.where(s>=72,'HOLD_OR_ADD_IF_THESIS_VALID','REASSESS')
 df['T3_3_6m']=np.where(df.score_medium_term>=65,'HOLD','REVIEW')
 df['T4_6_12m']=np.where(df.score_long_term>=65,'HOLD','REVIEW')
 df['T5_12_24m']=np.where(df.score_long_term>=70,'CORE_CANDIDATE','REASSESS')
 df['timing_reliable']=reliable
 return df

def rule(c):
 d={'score_short_term':'35% momentum + 25% analystes + 15% risque + 15% structure + 10% qualité','score_medium_term':'25% qualité + 20% value + 20% momentum + 20% analystes + 15% risque','score_long_term':'35% qualité + 25% value + 15% structure + 15% risque + 10% analystes','short_thesis_score':'45% momentum baissier + 30% faiblesse qualité + 15% analystes + 10% risque','committee_score_1829':'60% score absolu existant + 40% percentile multi-horizon','decision':'BUY_CANDIDATE si >77 et identité >=0.92; sinon WATCH/REVIEW/REJECT','execution':'RESEARCH_ONLY uniquement','T1_entry_low':'cours - 0.35 ATR si fiable','T1_entry_high':'cours + 0.10 ATR si fiable','T1_target':'objectif analystes si fiable','T1_invalidation':'source sinon cours - 1.6 ATR si fiable'}
 if c in d:return d[c]
 for fam,spec in FAMILIES.items():
  for src,h,w in spec:
   if c==src:return f'{fam}: poids {w}; percentile {"croissant" if h else "décroissant"}'
 return 'Critère source/provenance/contexte; non inventé.'

def comment(r):
 pos=[]; lim=[]
 for label,c in [('qualité','score_quality_100'),('momentum','score_momentum_100'),('analystes','score_analyst_100'),('valorisation','score_value_100')]:
  v=float(r.get(c,50) or 50)
  if v>=65:pos.append(label)
  elif v<40:lim.append(label)
 t='Qualifie par '+(', '.join(pos[:3]) if pos else 'un profil équilibré')+'.'
 if lim:t+=' Conviction limitée par '+', '.join(lim[:2])+'.'
 if not bool(r.get('timing_reliable')):t+=' Timing/prix non publiés faute de fiabilité suffisante.'
 return t

def run(root=None):
 root=root or ROOT; p=root/MASTER.relative_to(ROOT)
 df=pd.read_csv(p,sep=';',dtype=object,encoding='utf-8-sig')
 if len(df)!=1829 or df['isin'].astype(str).nunique()!=1829:raise RuntimeError('Final 1829 quality gate failed')
 src_cols=list(df.columns); df=timing(score(df.copy())); df['committee_comment']=df.apply(comment,axis=1)
 out=root/OUT.relative_to(ROOT); out.parent.mkdir(parents=True,exist_ok=True)
 df.sort_values('committee_score_1829',ascending=False).to_csv(out,sep=';',index=False,encoding='utf-8-sig')
 selected=df[pd.to_numeric(df.committee_score_1829,errors='coerce')>77].sort_values('committee_score_1829',ascending=False)
 wb=Workbook(); ws=wb.active; ws.title='Actions_gt_77'; cols=list(df.columns)
 for j,c in enumerate(cols,1):ws.cell(1,j,rule(c));ws.cell(2,j,c)
 for i,(_,r) in enumerate(selected.iterrows(),3):
  for j,c in enumerate(cols,1):
   v=r.get(c); ws.cell(i,j,None if pd.isna(v) else v)
 ws.freeze_panes='A3'; ws.auto_filter.ref=f'A2:{get_column_letter(len(cols))}{max(2,ws.max_row)}'
 for cell in ws[1]:cell.font=Font(bold=True);cell.fill=PatternFill('solid',fgColor='D9EAF7');cell.alignment=Alignment(wrap_text=True)
 for cell in ws[2]:cell.font=Font(bold=True);cell.fill=PatternFill('solid',fgColor='BDD7EE');cell.alignment=Alignment(wrap_text=True)
 gl=wb.create_sheet('Glossaire');gl.append(['Critère','Définition / transformation / rôle'])
 for c in cols:gl.append([c,rule(c)])
 audit={'rows':len(df),'source_columns':len(src_cols),'output_columns':len(df.columns),'selected_above_77':len(selected),'unique_isin':df['isin'].astype(str).nunique(),'score_min':float(pd.to_numeric(df.committee_score_1829).min()),'score_max':float(pd.to_numeric(df.committee_score_1829).max()),'smart_money_enabled':False,'live_order_execution_enabled':False,'execution_values':sorted(df.execution.astype(str).unique().tolist()),'generated_at':datetime.now(timezone.utc).isoformat(),'passed':bool(len(df)==1829 and df['isin'].astype(str).nunique()==1829 and df.execution.eq('RESEARCH_ONLY').all())}
 meta=wb.create_sheet('Metadata_Audit');meta.append(['Champ','Valeur'])
 for k,v in audit.items():meta.append([k,json.dumps(v,ensure_ascii=False) if isinstance(v,(dict,list)) else v])
 pack=root/PACK.relative_to(ROOT);pack.parent.mkdir(parents=True,exist_ok=True);wb.save(pack)
 ap=root/AUDIT.relative_to(ROOT);ap.parent.mkdir(parents=True,exist_ok=True);ap.write_text(json.dumps(audit,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
 print('V20_4_ACTIONS_1829_COMMITTEE',audit);return audit

if __name__=='__main__':run()
