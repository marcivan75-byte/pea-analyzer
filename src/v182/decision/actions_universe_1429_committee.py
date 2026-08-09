from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import json
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
MASTER = ROOT / "outputs" / "V20.4_GITOK_ACTIONS_1429_MASTER_ENRICHED.csv"
OUT = ROOT / "outputs" / "V20.4_GITOK_ACTIONS_1429_DECISIONS.csv"
PACK = ROOT / "outputs" / "V20.4_GITOK_ACTIONS_1429_ABOVE_77.xlsx"
AUDIT = ROOT / "outputs" / "audit" / "V20.4_ACTIONS_1429_COMMITTEE_AUDIT.json"

FAMILY_SPECS = {
    "quality": [("roe", True, 1.0), ("roa", True, 0.7), ("fcf_yield", True, 1.0), ("marge_ebit", True, 0.8), ("marge_nette", True, 0.7), ("croiss_ca_3y", True, 0.7), ("croiss_eps_3y", True, 0.8), ("dette_ebitda", False, 0.8), ("debt_to_equity", False, 0.5)],
    "value": [("per_forward", False, 1.0), ("per_ttm", False, 0.8), ("pb", False, 0.5), ("ev_ebit", False, 0.8), ("fcf_yield", True, 1.0), ("per_vs_sector_pct", False, 0.6), ("target_upside_pct", True, 0.6)],
    "momentum": [("perf_1m_pct", True, 1.0), ("perf_3m_pct", True, 1.2), ("perf_6m_pct", True, 1.2), ("perf_1y_pct", True, 0.7), ("relative_strength", True, 1.0), ("macd_hist", True, 0.7), ("rsi14", True, 0.3), ("rvol20", True, 0.4)],
    "analyst": [("analyst_momentum_score", True, 1.2), ("consensus_score_100", True, 1.0), ("target_upside_pct", True, 1.0), ("weighted_target_revision_30d_pct", True, 0.8), ("weighted_consensus_delta_30d", True, 0.8), ("revision_breadth_30d", True, 0.7), ("net_upgrades_30d", True, 0.6), ("consensus_confidence", True, 0.4)],
    "risk": [("volatility_20d", False, 1.0), ("volatility_60d", False, 0.8), ("max_drawdown_1y", True, 1.0), ("beta", False, 0.5), ("asymmetry", True, 0.8), ("data_trust_pct", True, 0.5)],
    "structure": [("market_cap", True, 0.8), ("volume", True, 0.6), ("coverage_pct", True, 0.7), ("data_trust_pct", True, 0.7), ("v182_ticker_validation_confidence_pct", True, 0.4)],
}


def _num(df: pd.DataFrame, col: str) -> pd.Series:
    if col not in df.columns:
        return pd.Series(np.nan, index=df.index, dtype=float)
    return pd.to_numeric(df[col], errors="coerce")


def _rank(df: pd.DataFrame, col: str, higher: bool) -> pd.Series:
    s = _num(df, col)
    p = s.rank(pct=True, method="average") * 100.0
    if not higher:
        p = 100.0 - p
    return p.fillna(50.0).clip(0, 100)


def _family(df: pd.DataFrame, specs: list[tuple[str, bool, float]]) -> pd.Series:
    vals, weights = [], []
    for col, higher, weight in specs:
        if col in df.columns:
            vals.append(_rank(df, col, higher) * weight)
            weights.append(weight)
    return sum(vals) / sum(weights) if vals else pd.Series(50.0, index=df.index)


def _score(df: pd.DataFrame) -> pd.DataFrame:
    fam = {k: _family(df, v) for k, v in FAMILY_SPECS.items()}
    for k, v in fam.items():
        df[f"score_{k}_100"] = v.round(2)
    df["score_short_term"] = (0.35*fam["momentum"] + 0.25*fam["analyst"] + 0.15*fam["risk"] + 0.15*fam["structure"] + 0.10*fam["quality"]).round(2)
    df["score_medium_term"] = (0.25*fam["quality"] + 0.20*fam["value"] + 0.20*fam["momentum"] + 0.20*fam["analyst"] + 0.15*fam["risk"]).round(2)
    df["score_long_term"] = (0.35*fam["quality"] + 0.25*fam["value"] + 0.15*fam["structure"] + 0.15*fam["risk"] + 0.10*fam["analyst"]).round(2)
    df["short_thesis_score"] = (0.45*(100-fam["momentum"]) + 0.30*(100-fam["quality"]) + 0.15*(100-fam["analyst"]) + 0.10*(100-fam["risk"])).round(2)
    multi = 0.25*df["score_short_term"] + 0.35*df["score_medium_term"] + 0.40*df["score_long_term"]
    base = pd.to_numeric(df.get("committee_score_with_analyst_momentum"), errors="coerce")
    fallback = pd.to_numeric(df.get("score_brut"), errors="coerce")
    base = base.fillna(fallback).fillna(50.0).clip(0, 100)
    percentile = multi.rank(pct=True, method="average") * 100.0
    df["committee_score_1429"] = (0.60*base + 0.40*percentile).round(2)
    return df


def _timing(df: pd.DataFrame) -> pd.DataFrame:
    conf = _num(df, "v182_ticker_validation_confidence_pct").fillna(0) / 100.0
    trust = _num(df, "data_trust_pct").fillna(0) / 100.0
    last, atr = _num(df, "last_close"), _num(df, "atr14")
    target = _num(df, "target_price").fillna(_num(df, "target_mean_yf"))
    inval = _num(df, "invalidation_level")
    reliable = (conf >= 0.92) & (trust >= 0.50) & last.notna() & (last > 0)
    df["identity_method"] = np.where(conf >= 0.92, "CANONICAL_TICKER_ISIN", "CANONICAL_ISIN")
    df["identity_confidence"] = conf.round(3)
    df["T1_entry_low"] = np.where(reliable & atr.notna(), (last - 0.35*atr).clip(lower=0), np.nan)
    df["T1_entry_high"] = np.where(reliable & atr.notna(), last + 0.10*atr, np.nan)
    df["T1_target"] = np.where(reliable & target.notna(), target, np.nan)
    df["T1_invalidation"] = np.where(reliable & inval.notna(), inval, np.where(reliable & atr.notna(), (last - 1.6*atr).clip(lower=0), np.nan))
    score = df["committee_score_1429"]
    df["decision"] = np.select([score > 77, score >= 70, score >= 60], ["BUY_CANDIDATE", "WATCH", "REVIEW"], default="REJECT")
    low = conf < 0.92
    df.loc[low, "decision"] = np.where(score[low] >= 70, "REVIEW", "REJECT")
    df["execution"] = "RESEARCH_ONLY"
    df["T0"] = np.select([df["decision"].eq("BUY_CANDIDATE"), df["decision"].eq("WATCH")], ["PREPARE", "WATCH"], default="NO_ACTION")
    df["T1_1_4w"] = np.where(df["decision"].eq("BUY_CANDIDATE"), "ENTRY_IF_ZONE_VALID", "MONITOR")
    df["T2_1_3m"] = np.where(score >= 72, "HOLD_OR_ADD_IF_THESIS_VALID", "REASSESS")
    df["T3_3_6m"] = np.where(df["score_medium_term"] >= 65, "HOLD", "REVIEW")
    df["T4_6_12m"] = np.where(df["score_long_term"] >= 65, "HOLD", "REVIEW")
    df["T5_12_24m"] = np.where(df["score_long_term"] >= 70, "CORE_CANDIDATE", "REASSESS")
    df["timing_reliable"] = reliable
    return df


def _rule(col: str) -> str:
    derived = {
        "score_short_term": "35% momentum + 25% analystes + 15% risque + 15% structure + 10% qualité",
        "score_medium_term": "25% qualité + 20% value + 20% momentum + 20% analystes + 15% risque",
        "score_long_term": "35% qualité + 25% value + 15% structure + 15% risque + 10% analystes",
        "short_thesis_score": "45% momentum baissier + 30% faiblesse qualité + 15% analystes + 10% risque",
        "committee_score_1429": "60% comité absolu existant + 40% rang percentile multi-horizon",
        "decision": "BUY_CANDIDATE si score >77 et identité >=0.92; sinon WATCH/REVIEW/REJECT",
        "execution": "Toujours RESEARCH_ONLY; jamais d'ordre réel",
        "T1_entry_low": "Dernier cours - 0.35 ATR si données fiables",
        "T1_entry_high": "Dernier cours + 0.10 ATR si données fiables",
        "T1_target": "Objectif analystes canonique/Yahoo si fiable",
        "T1_invalidation": "Niveau source sinon dernier cours - 1.6 ATR si fiable",
    }
    if col in derived:
        return derived[col]
    for fam, specs in FAMILY_SPECS.items():
        for source, higher, weight in specs:
            if col == source:
                return f"{fam}: poids {weight}; percentile {'croissant' if higher else 'décroissant'}"
    return "Critère source canonique / provenance / contexte; non transformé directement sauf mention"


def _comment(row: pd.Series) -> str:
    positives, limits = [], []
    for label, col in [("qualité", "score_quality_100"), ("momentum", "score_momentum_100"), ("analystes", "score_analyst_100"), ("valorisation", "score_value_100")]:
        value = float(row.get(col, 50) or 50)
        if value >= 65:
            positives.append(label)
        elif value < 40:
            limits.append(label)
    text = "Qualifie par " + (", ".join(positives[:3]) if positives else "un profil équilibré") + "."
    if limits:
        text += " Conviction limitée par " + ", ".join(limits[:2]) + "."
    if not bool(row.get("timing_reliable")):
        text += " Timing/prix non publiés faute de fiabilité suffisante."
    return text


def run(root: Path | None = None) -> dict:
    root = root or ROOT
    master = root / MASTER.relative_to(ROOT)
    df = pd.read_csv(master, sep=";", dtype=object, encoding="utf-8-sig")
    if len(df) != 1429 or df["isin"].astype(str).nunique() != 1429:
        raise RuntimeError("Canonical 1429 quality gate failed")
    source_columns = list(df.columns)
    df = _timing(_score(df.copy()))
    df["committee_comment"] = df.apply(_comment, axis=1)
    out = root / OUT.relative_to(ROOT)
    out.parent.mkdir(parents=True, exist_ok=True)
    df.sort_values("committee_score_1429", ascending=False).to_csv(out, sep=";", index=False, encoding="utf-8-sig")
    selected = df.loc[pd.to_numeric(df["committee_score_1429"], errors="coerce") > 77].sort_values("committee_score_1429", ascending=False)
    wb = Workbook()
    ws = wb.active
    ws.title = "Actions_gt_77"
    cols = list(df.columns)
    for j, col in enumerate(cols, 1):
        ws.cell(1, j, _rule(col))
        ws.cell(2, j, col)
    for i, (_, row) in enumerate(selected.iterrows(), 3):
        for j, col in enumerate(cols, 1):
            value = row.get(col)
            ws.cell(i, j, None if pd.isna(value) else value)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{max(2, ws.max_row)}"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="BDD7EE")
        cell.alignment = Alignment(wrap_text=True)
    glossary = wb.create_sheet("Glossaire")
    glossary.append(["Critère", "Définition / transformation / rôle"])
    for col in cols:
        glossary.append([col, _rule(col)])
    audit = {
        "rows": int(len(df)),
        "source_columns": int(len(source_columns)),
        "output_columns": int(len(df.columns)),
        "selected_above_77": int(len(selected)),
        "unique_isin": int(df["isin"].astype(str).nunique()),
        "score_min": float(pd.to_numeric(df["committee_score_1429"]).min()),
        "score_max": float(pd.to_numeric(df["committee_score_1429"]).max()),
        "smart_money_enabled": False,
        "live_order_execution_enabled": False,
        "execution_values": sorted(df["execution"].astype(str).unique().tolist()),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "passed": bool(len(df) == 1429 and df["isin"].astype(str).nunique() == 1429 and df["execution"].eq("RESEARCH_ONLY").all()),
    }
    metadata = wb.create_sheet("Metadata_Audit")
    metadata.append(["Champ", "Valeur"])
    for key, value in audit.items():
        metadata.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    pack = root / PACK.relative_to(ROOT)
    pack.parent.mkdir(parents=True, exist_ok=True)
    wb.save(pack)
    audit_path = root / AUDIT.relative_to(ROOT)
    audit_path.parent.mkdir(parents=True, exist_ok=True)
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return audit


def main() -> None:
    print("V20_4_ACTIONS_1429_COMMITTEE", run())


if __name__ == "__main__":
    main()
