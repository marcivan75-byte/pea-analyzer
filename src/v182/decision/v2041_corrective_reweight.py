from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import json
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parents[3]
A_IN = ROOT / 'outputs/V20.4_GITOK_ACTIONS_1829_DECISIONS.csv'
E_IN = ROOT / 'outputs/V20.4_GITOK_ETF_266_DECISIONS.csv'
A_OUT = ROOT / 'outputs/V20.4.1_GITOK_ACTIONS_1829_DECISIONS.csv'
E_OUT = ROOT / 'outputs/V20.4.1_GITOK_ETF_266_DECISIONS.csv'
PACK = ROOT / 'outputs/V20.4.1_COMMITTEE_COMPARISON.xlsx'
AUDIT = ROOT / 'outputs/audit/V20.4.1_CORRECTIVE_AUDIT.json'

A_FAMILIES = {
    'quality': [('roe', True, 1.0), ('roa', True, .7), ('marge_ebit', True, .8), ('marge_nette', True, .7), ('croiss_ca_3y', True, .7), ('croiss_eps_3y', True, .8), ('dette_ebitda', False, .8), ('debt_to_equity', False, .5)],
    'value': [('per_forward', False, 1.0), ('per_ttm', False, .8), ('pb', False, .5), ('ev_ebit', False, .8), ('fcf_yield', True, 1.0), ('per_vs_sector_pct', False, .6)],
    'momentum': [('perf_1m_pct', True, 1.0), ('perf_3m_pct', True, 1.15), ('perf_6m_pct', True, 1.10), ('perf_1y_pct', True, .5), ('relative_strength', True, 1.0), ('macd_hist', True, .8), ('RSI_BAND', True, .45), ('rvol20', True, .6)],
    'analyst': [('analyst_momentum_score', True, 1.2), ('consensus_score_100', True, 1.0), ('target_upside_pct', True, .9), ('weighted_target_revision_30d_pct', True, .9), ('weighted_consensus_delta_30d', True, .9), ('revision_breadth_30d', True, .8), ('net_upgrades_30d', True, .7), ('consensus_confidence', True, .5)],
    'risk': [('volatility_20d', False, 1.0), ('volatility_60d', False, .8), ('max_drawdown_1y', True, 1.1), ('beta', False, .5), ('asymmetry', True, 1.0)],
    'structure': [('market_cap', True, .9), ('volume', True, .8)],
}

A_HORIZONS = {
    'short': {'quality': .10, 'value': .03, 'momentum': .32, 'analyst': .23, 'risk': .20, 'structure': .12},
    'medium': {'quality': .25, 'value': .19, 'momentum': .19, 'analyst': .18, 'risk': .19, 'structure': 0.0},
    'long': {'quality': .35, 'value': .24, 'momentum': .03, 'analyst': .08, 'risk': .20, 'structure': .10},
}


def _num(df: pd.DataFrame, col: str) -> pd.Series:
    if col not in df.columns:
        return pd.Series(np.nan, index=df.index, dtype=float)
    return pd.to_numeric(df[col], errors='coerce')


def _rank(df: pd.DataFrame, col: str, higher: bool) -> pd.Series:
    p = _num(df, col).rank(pct=True, method='average') * 100
    return (p if higher else 100 - p).fillna(50).clip(0, 100)


def _rsi_band(df: pd.DataFrame) -> pd.Series:
    # Corrects the V20.4 monotonic RSI error: high RSI is not indefinitely better.
    rsi = _num(df, 'rsi14')
    score = (100 - 2.5 * (rsi - 60).abs()).clip(0, 100)
    return score.fillna(50)


def _family(df: pd.DataFrame, spec: list[tuple[str, bool, float]]) -> pd.Series:
    vals, weights = [], []
    for col, higher, weight in spec:
        value = _rsi_band(df) if col == 'RSI_BAND' else _rank(df, col, higher)
        vals.append(value * weight)
        weights.append(weight)
    return sum(vals) / sum(weights)


def _score_actions(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    fam = {name: _family(out, spec) for name, spec in A_FAMILIES.items()}
    for name, values in fam.items():
        out[f'v2041_score_{name}_100'] = values.round(2)

    def horizon(weights: dict[str, float]) -> pd.Series:
        return sum(fam[name] * weight for name, weight in weights.items())

    out['v2041_score_short_term'] = horizon(A_HORIZONS['short']).round(2)
    out['v2041_score_medium_term'] = horizon(A_HORIZONS['medium']).round(2)
    out['v2041_score_long_term'] = horizon(A_HORIZONS['long']).round(2)
    out['v2041_short_thesis_score'] = (
        .35 * (100 - fam['momentum']) + .25 * (100 - fam['quality']) + .15 * (100 - fam['value'])
        + .15 * (100 - fam['analyst']) + .10 * (100 - fam['risk'])
    ).round(2)

    multi = .25 * out['v2041_score_short_term'] + .35 * out['v2041_score_medium_term'] + .40 * out['v2041_score_long_term']
    pct = multi.rank(pct=True, method='average') * 100
    base = _num(out, 'committee_score_with_analyst_momentum').fillna(_num(out, 'score_brut')).fillna(50).clip(0, 100)

    # Data quality ceases to be alpha: only a small fail-soft penalty applies to weak evidence.
    trust = _num(out, 'data_trust_pct').fillna(50).clip(0, 100)
    coverage = _num(out, 'coverage_pct').fillna(50).clip(0, 100)
    evidence = np.minimum(trust, coverage)
    confidence_factor = np.where(evidence < 40, .96, np.where(evidence < 55, .985, 1.0))

    out['v2041_confidence_factor'] = pd.Series(confidence_factor, index=out.index).round(3)
    out['v2041_committee_score'] = ((.60 * base + .40 * pct) * confidence_factor).clip(0, 100).round(2)
    out['v204_decision'] = out.get('decision', 'REJECT').astype(str)
    score = out['v2041_committee_score']
    out['decision'] = np.select([score > 77, score >= 70, score >= 60], ['BUY_CANDIDATE', 'WATCH', 'REVIEW'], default='REJECT')
    identity = _num(out, 'identity_confidence').fillna(0)
    low = identity < .92
    out.loc[low, 'decision'] = np.where(score[low] >= 70, 'REVIEW', 'REJECT')
    out['execution'] = 'RESEARCH_ONLY'
    out['v2041_version'] = 'V20.4.1_CORRECTIVE'
    return out


def _rank_series(series: pd.Series, higher: bool = True) -> pd.Series:
    s = pd.to_numeric(series, errors='coerce')
    p = s.rank(pct=True, method='average') * 100
    return (p if higher else 100 - p).fillna(50)


def _ter_score(series: pd.Series) -> pd.Series:
    x = pd.to_numeric(series, errors='coerce')
    return (100 * (1 - (x - .15) / .85)).clip(0, 100).fillna(50)


def _score_etf(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    base = pd.to_numeric(out.get('Score V9'), errors='coerce').fillna(pd.to_numeric(out.get('SCORE V5 /100'), errors='coerce')).fillna(50)
    liq = pd.to_numeric(out.get('Score liquidité combiné /100'), errors='coerce').fillna(50)
    rep = pd.to_numeric(out.get('Score réplication /100'), errors='coerce').fillna(50)
    esg = pd.to_numeric(out.get('Score ESG détaillé /100'), errors='coerce').fillna(50)
    ter = _ter_score(out.get('TER %', pd.Series(index=out.index, dtype=float)))
    diversification = (
        _rank_series(out.get('Nombre holdings', pd.Series(index=out.index, dtype=float)), True)
        + _rank_series(out.get('Top 10 holdings %', pd.Series(index=out.index, dtype=float)), False)
        + _rank_series(out.get('HHI holdings', pd.Series(index=out.index, dtype=float)), False)
    ) / 3
    risk_efficiency = (
        _rank_series(out.get('Sharpe', pd.Series(index=out.index, dtype=float)), True)
        + _rank_series(out.get('Max DD %', pd.Series(index=out.index, dtype=float)), True)
        + _rank_series(out.get('Vol 5a %', pd.Series(index=out.index, dtype=float)), False)
    ) / 3

    # Removes duplicated AUM/spread/ADV alpha adjustments: those remain hard/liquidity gates, not extra alpha.
    structural = .52 * base + .16 * liq + .10 * rep + .03 * esg + .07 * ter + .05 * diversification + .07 * risk_efficiency
    technical = pd.to_numeric(out.get('technical_adjustment'), errors='coerce').fillna(0)
    availability = (pd.to_numeric(out.get('Dispo Score /3'), errors='coerce').fillna(0).clip(0, 3) / 3)
    availability_factor = .97 + .03 * availability

    out['v2041_structural_score_266'] = structural.clip(0, 100).round(2)
    out['v2041_availability_factor'] = availability_factor.round(3)
    out['v2041_etf_266_score'] = ((structural + technical) * availability_factor).clip(0, 100).round(2)
    out['v204_decision'] = out.get('decision', 'REJECT').astype(str)

    aum = pd.to_numeric(out.get('AUM M€'), errors='coerce')
    spread = pd.to_numeric(out.get('Spread %'), errors='coerce')
    daily = out.get('daily_match', False).astype(str).str.lower().isin({'true', '1', 'yes', 'oui'})
    conf = pd.to_numeric(out.get('daily_match_confidence'), errors='coerce').fillna(0)
    p3 = pd.to_numeric(out.get('daily_perf_3m_pct'), errors='coerce').fillna(0)
    p1y = pd.to_numeric(out.get('daily_perf_1y_pct'), errors='coerce').fillna(0)
    rsi = pd.to_numeric(out.get('daily_rsi14'), errors='coerce')
    score = out['v2041_etf_266_score']
    structural_score = out['v2041_structural_score_266']

    decisions = []
    reasons = []
    executions = []
    for i in out.index:
        hard = (pd.notna(aum.loc[i]) and aum.loc[i] < 15) or (pd.notna(spread.loc[i]) and spread.loc[i] > 1.50)
        if hard:
            d, reason, exe = 'REJECT', 'ETF_HARD_LIQUIDITY_GATE', 'RESEARCH_ONLY'
        elif daily.loc[i] and conf.loc[i] >= .95 and score.loc[i] >= 72 and p3.loc[i] > 0 and p1y.loc[i] > 0 and (pd.isna(rsi.loc[i]) or rsi.loc[i] <= 72):
            d, reason, exe = 'BUY_CANDIDATE', 'V2041_STRUCTURE_AND_TIMING_CONFIRMED', 'RECOMMENDATION_ONLY'
        elif daily.loc[i] and score.loc[i] >= 65:
            d, reason, exe = 'WATCH', 'V2041_STRUCTURE_OK_TIMING_PARTIAL', 'RECOMMENDATION_ONLY'
        elif daily.loc[i]:
            d, reason, exe = 'REVIEW', 'V2041_DAILY_SIGNAL_MIXED', 'RESEARCH_ONLY'
        elif structural_score.loc[i] >= 72:
            d, reason, exe = 'STRUCTURAL_CANDIDATE', 'V2041_STRONG_STRUCTURE_DAILY_DATA_REQUIRED', 'RESEARCH_ONLY'
        elif structural_score.loc[i] >= 55:
            d, reason, exe = 'DATA_REQUIRED', 'V2041_STRUCTURE_NOT_REJECTED_DAILY_DATA_REQUIRED', 'RESEARCH_ONLY'
        else:
            d, reason, exe = 'REJECT', 'V2041_STRUCTURAL_QUALITY_INSUFFICIENT', 'RESEARCH_ONLY'
        decisions.append(d); reasons.append(reason); executions.append(exe)

    out['decision'] = decisions
    out['decision_reason'] = reasons
    out['execution'] = executions
    out['v2041_version'] = 'V20.4.1_CORRECTIVE'
    return out


def _median(frame: pd.DataFrame, col: str, mask: pd.Series) -> float | None:
    if col not in frame.columns:
        return None
    v = pd.to_numeric(frame.loc[mask, col], errors='coerce').median()
    return None if pd.isna(v) else float(v)


def _write_pack(actions: pd.DataFrame, etfs: pd.DataFrame, audit: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Synthese'
    rows = [
        ['V20.4.1 corrective', 'Actions PEA', 'ETF PEA'],
        ['Lignes', len(actions), len(etfs)],
        ['BUY V20.4', int((actions.v204_decision == 'BUY_CANDIDATE').sum()), int((etfs.v204_decision == 'BUY_CANDIDATE').sum())],
        ['BUY V20.4.1', int((actions.decision == 'BUY_CANDIDATE').sum()), int((etfs.decision == 'BUY_CANDIDATE').sum())],
        ['Smart Money', 'NON INTEGRE', 'NON INTEGRE'],
        ['Live/order execution', 'DISABLED', 'DISABLED'],
    ]
    for row in rows: ws.append(row)
    for c in ws[1]: c.font = Font(bold=True); c.fill = PatternFill('solid', fgColor='17365D'); c.font = Font(bold=True, color='FFFFFF')

    w = wb.create_sheet('Ponderations')
    w.append(['Univers', 'Bloc', 'V20.4.1'])
    w.append(['Actions', 'Court terme', '10% Qualite + 3% Valorisation + 32% Momentum + 23% Analystes + 20% Risque + 12% Structure'])
    w.append(['Actions', 'Moyen terme', '25% Qualite + 19% Valorisation + 19% Momentum + 18% Analystes + 19% Risque'])
    w.append(['Actions', 'Long terme', '35% Qualite + 24% Valorisation + 3% Momentum + 8% Analystes + 20% Risque + 10% Structure'])
    w.append(['Actions', 'Score final', '60% score absolu amont + 40% percentile multi-horizon; penalite evidence seulement si faible'])
    w.append(['Actions', 'Corrections', 'RSI non monotone; target_upside retire de Value; FCF yield retire de Quality; data trust/coverage/ticker confidence ne produisent plus alpha direct'])
    w.append(['ETF', 'Structure', '52% Score V9 +16% liquidite +10% replication +3% ESG +7% TER +5% diversification +7% efficacite risque'])
    w.append(['ETF', 'Corrections', 'Disponibilite devient facteur de confiance; AUM/spread/ADV restent gates/liquidite mais plus de double bonus alpha'])
    for c in w[1]: c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='17365D')
    w.column_dimensions['A'].width = 18; w.column_dimensions['B'].width = 26; w.column_dimensions['C'].width = 105
    for row in w.iter_rows():
        for cell in row: cell.alignment = Alignment(wrap_text=True, vertical='top')

    cmp = wb.create_sheet('Actions_BUY_compare')
    old_buy = actions.v204_decision.eq('BUY_CANDIDATE')
    new_buy = actions.decision.eq('BUY_CANDIDATE')
    cmp.append(['Metrique', 'V20.4 BUY', 'V20.4.1 BUY'])
    for label, col in [('Volatilite 20j mediane','volatility_20d'),('Volatilite 60j mediane','volatility_60d'),('Max drawdown 1a median','max_drawdown_1y'),('Target upside median %','target_upside_pct'),('Data trust median %','data_trust_pct')]:
        cmp.append([label, _median(actions,col,old_buy), _median(actions,col,new_buy)])
    for c in cmp[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='17365D')

    meta = wb.create_sheet('Audit')
    meta.append(['Champ','Valeur'])
    for k,v in audit.items(): meta.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v,(dict,list)) else v])
    for c in meta[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='17365D')
    meta.column_dimensions['A'].width=42; meta.column_dimensions['B'].width=105
    wb.save(PACK)


def run() -> dict:
    if not A_IN.exists() or not E_IN.exists():
        raise RuntimeError('V20.4 decision inputs missing')
    actions0 = pd.read_csv(A_IN, sep=';', encoding='utf-8-sig', low_memory=False)
    etfs0 = pd.read_csv(E_IN, sep=';', encoding='utf-8-sig', low_memory=False)
    if len(actions0) != 1829 or actions0['isin'].astype(str).nunique() != 1829:
        raise RuntimeError('Actions 1829 input gate failed')
    if len(etfs0) != 266 or etfs0['ISIN'].astype(str).nunique() != 266:
        raise RuntimeError('ETF 266 input gate failed')

    actions = _score_actions(actions0)
    etfs = _score_etf(etfs0)
    A_OUT.parent.mkdir(parents=True, exist_ok=True)
    actions.to_csv(A_OUT, sep=';', index=False, encoding='utf-8-sig')
    etfs.to_csv(E_OUT, sep=';', index=False, encoding='utf-8-sig')

    live = {'LIVE','ORDER','EXECUTE','BROKER','REAL_ORDER','LIVE_ORDER'}
    passed = (
        len(actions) == 1829 and actions['isin'].astype(str).nunique() == 1829
        and len(etfs) == 266 and etfs['ISIN'].astype(str).nunique() == 266
        and not actions.execution.astype(str).str.upper().isin(live).any()
        and not etfs.execution.astype(str).str.upper().isin(live).any()
        and not ((actions.decision == 'BUY_CANDIDATE') & (_num(actions,'identity_confidence').fillna(0) < .92)).any()
    )
    audit = {
        'version': 'V20.4.1_CORRECTIVE',
        'generated_at': datetime.now(timezone.utc).isoformat(),
        'passed': bool(passed),
        'actions_rows': len(actions),
        'etf_rows': len(etfs),
        'actions_v204_decisions': actions.v204_decision.value_counts().to_dict(),
        'actions_v2041_decisions': actions.decision.value_counts().to_dict(),
        'etf_v204_decisions': etfs.v204_decision.value_counts().to_dict(),
        'etf_v2041_decisions': etfs.decision.value_counts().to_dict(),
        'actions_buy_overlap': int(((actions.v204_decision == 'BUY_CANDIDATE') & (actions.decision == 'BUY_CANDIDATE')).sum()),
        'smart_money_integrated': False,
        'live_order_execution_enabled': False,
        'corrections': [
            'RSI band scoring replaces monotonic higher-is-better RSI',
            'target_upside removed from Value family to avoid duplicate analyst alpha',
            'fcf_yield kept in Value only, removed from Quality',
            'data_trust/coverage/ticker-confidence cease to be direct alpha; evidence only penalizes weak data',
            'risk family weight increased across horizons',
            'small valuation added to short-term and small momentum added to long-term',
            'ETF availability converted from alpha to confidence factor',
            'ETF duplicated AUM/spread/ADV alpha adjustments removed; hard liquidity gates retained',
        ],
    }
    AUDIT.parent.mkdir(parents=True, exist_ok=True)
    AUDIT.write_text(json.dumps(audit, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    _write_pack(actions, etfs, audit)
    if not passed:
        raise RuntimeError('V20.4.1 quality gate failed')
    print('V20_4_1_OK', audit)
    return audit


if __name__ == '__main__':
    run()
