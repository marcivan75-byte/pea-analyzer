from pathlib import Path

from docx import Document
from openpyxl import load_workbook
import pandas as pd

from v182.reporting.committee_ci_explainability import (
    REFERENCE_COLUMNS,
    _decision_comment,
    _generic_details,
    _write_excel_reference,
    _write_word_report,
    run,
)


def test_generic_details_reconstructs_cross_sectional_score_exactly():
    source = pd.DataFrame(
        [
            {"isin": "FR0001", "name": "A", "c1": 10.0, "c2": 5.0},
            {"isin": "FR0002", "name": "B", "c1": 20.0, "c2": 1.0},
        ]
    )
    selected = pd.DataFrame(
        [
            {"asset_class": "ACTION", "horizon": "CT", "isin": "FR0001", "name": "A", "decision": "BUY_CANDIDATE", "score": 75.0},
        ]
    )
    registry = {
        "weights": {"CT": {"c1": 0.5, "c2": 0.5}},
        "directions": {"CT": {"c1": "HIGH", "c2": "HIGH"}},
        "horizons": {"CT": {"minimum_weighted_coverage": 0.7}},
    }
    detail = _generic_details(source, selected, registry, "ACTION", ["CT"])
    assert len(detail) == 2
    assert abs(detail["effective_weight_pct"].sum() - 100.0) < 1e-9
    assert abs(detail["weighted_contribution_points"].sum() - 75.0) < 1e-9


def test_missing_criterion_is_explicit_and_not_renormalized_as_available():
    source = pd.DataFrame(
        [
            {"isin": "FR0001", "name": "A", "c1": 10.0},
            {"isin": "FR0002", "name": "B", "c1": 20.0},
        ]
    )
    selected = pd.DataFrame(
        [{"asset_class": "ACTION", "horizon": "CT", "isin": "FR0001", "name": "A", "decision": "WATCH", "score": 50.0}]
    )
    registry = {
        "weights": {"CT": {"c1": 0.5, "missing": 0.5}},
        "directions": {"CT": {"c1": "HIGH", "missing": "HIGH"}},
        "horizons": {"CT": {"minimum_weighted_coverage": 0.7}},
    }
    detail = _generic_details(source, selected, registry, "ACTION", ["CT"])
    missing = detail.loc[detail["criterion"] == "missing"].iloc[0]
    active = detail.loc[detail["criterion"] == "c1"].iloc[0]
    assert missing["criterion_status"] == "MISSING"
    assert missing["effective_weight_pct"] == 0.0
    assert active["effective_weight_pct"] == 100.0


def _ci_fixture() -> tuple[pd.DataFrame, pd.DataFrame]:
    context = pd.DataFrame(
        [
            {
                "asset_class": "ACTION",
                "horizon": "CT",
                "isin": "FR0001",
                "name": "Valeur A",
                "decision": "BUY_CANDIDATE",
                "score": 82.5,
                "coverage_pct": 91.0,
                "v21_8_entry_state": "WAIT_ENTRY",
                "v21_8_position_state": "NO_POSITION",
                "v21_8_entry_reasons": "confirmation requise",
                "v21_8_position_reasons": "aucune position",
                "risk_verdict": "CONTEXT_OK",
                "valuation_warning": None,
                "correction_alert": None,
                "notes": "Référence gelée",
            }
        ]
    )
    detail = pd.DataFrame(
        [
            {
                "asset_class": "ACTION",
                "horizon": "CT",
                "name": "Valeur A",
                "isin": "FR0001",
                "decision": "BUY_CANDIDATE",
                "final_score": 82.5,
                "criterion": "momentum",
                "criterion_status": "ACTIVE",
                "raw_value": 0.12,
                "direction": "HIGH",
                "criterion_score_0_100": 90.0,
                "theoretical_weight_pct": 60.0,
                "effective_weight_pct": 60.0,
                "weighted_contribution_points": 54.0,
                "source": "TEST_SOURCE",
                "as_of": "2026-08-22",
                "evidence_level": "PRIMARY",
                "validation_status": "VALID",
            },
            {
                "asset_class": "ACTION",
                "horizon": "CT",
                "name": "Valeur A",
                "isin": "FR0001",
                "decision": "BUY_CANDIDATE",
                "final_score": 82.5,
                "criterion": "risk",
                "criterion_status": "ACTIVE",
                "raw_value": 0.18,
                "direction": "LOW",
                "criterion_score_0_100": 71.25,
                "theoretical_weight_pct": 40.0,
                "effective_weight_pct": 40.0,
                "weighted_contribution_points": 28.5,
                "source": "TEST_SOURCE",
                "as_of": "2026-08-22",
                "evidence_level": "PRIMARY",
                "validation_status": "VALID",
            },
        ]
    )
    return context, detail


def test_decision_comment_states_score_coverage_and_reference_counts():
    context, detail = _ci_fixture()
    comment = _decision_comment(next(context.itertuples()), detail)
    assert "82.5/100" in comment
    assert "91.0%" in comment
    assert "2 critères actifs" in comment
    assert "sans recalcul ni modification du score" in comment


def test_word_report_is_openable_and_contains_argumented_sections(tmp_path: Path):
    context, detail = _ci_fixture()
    path = tmp_path / "CI_COMITE_INVESTISSEMENT.docx"
    _write_word_report(path, context, detail)
    assert path.exists()
    document = Document(path)
    text = "\n".join(paragraph.text for paragraph in document.paragraphs)
    assert "Analyse détaillée par valeur" in text
    assert "Facteurs les plus contributifs" in text
    assert "Critères les moins favorables" in text
    assert "Conclusion CI" in text
    assert "Valeur A" in text


def test_excel_contains_only_weighted_reference_sheet(tmp_path: Path):
    _, detail = _ci_fixture()
    path = tmp_path / "CI_REFERENTIEL_PONDERE.xlsx"
    _write_excel_reference(path, detail)
    workbook = load_workbook(path, read_only=True)
    assert workbook.sheetnames == ["Referentiel_pondere"]
    worksheet = workbook["Referentiel_pondere"]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    assert headers == REFERENCE_COLUMNS
    assert worksheet.max_row == 3


def test_run_fails_closed_without_committee_decisions(tmp_path: Path):
    result = run(tmp_path)
    assert result["status"] == "BLOCKED_COMMITTEE_DECISIONS_MISSING"
    assert result["real_orders_enabled"] is False
